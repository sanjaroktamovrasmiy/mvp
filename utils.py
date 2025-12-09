#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Yordamchi funksiyalar
"""

import logging
import os
import re
import shutil
import textwrap
import pdfkit
import numpy as np
from io import BytesIO
from datetime import datetime
from telegram import Update
from telegram.ext import ContextTypes
from database import load_data
from openpyxl import Workbook, load_workbook
from scipy.special import expit

logger = logging.getLogger(__name__)


async def check_subscription(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Majburiy kanallarga obuna tekshiruvi"""
    data = load_data()
    if not data["mandatory_channels"]:
        return True
    
    user_id = update.effective_user.id
    for channel in data["mandatory_channels"]:
        try:
            # Kanal username yoki ID bo'lishi mumkin
            channel_id = channel if channel.startswith('@') or channel.startswith('-') else f"@{channel}"
            member = await context.bot.get_chat_member(channel_id, user_id)
            if member.status in ['left', 'kicked']:
                return False
        except Exception as e:
            logger.error(f"Kanal tekshiruvi xatosi: {e} - Kanal: {channel}")
            # Agar kanal topilmasa, xavfsizlik uchun False qaytaramiz
            return False
    return True


def generate_pdf(result_id, result_data):
    """PDF fayl yaratish (pdfkit + reportlab fallback)"""
    html_content = None
    fallback_lines = result_data.get('fallback_lines')
    fallback_title = result_data.get('fallback_title') or result_data.get('test_name') or "Natijalar hisobot"
    
    try:
        if 'html_content' in result_data:
            html_content = result_data['html_content']
        else:
            # Oddiy foydalanuvchi natijasi uchun HTML yaratamiz
            html_content = _build_default_result_html(result_data)
            if fallback_lines is None:
                fallback_lines = _build_fallback_lines_from_result(result_data)
        
        # Avval pdfkit orqali urinib ko'ramiz (wkhtmltopdf talab qiladi)
        if html_content:
            try:
                config = None
                wkhtml_path = shutil.which("wkhtmltopdf")
                if wkhtml_path:
                    config = pdfkit.configuration(wkhtmltopdf=wkhtml_path)
                pdf_bytes = pdfkit.from_string(
                    html_content,
                    False,
                    configuration=config,
                    options={
                        'page-size': 'A4',
                        'encoding': 'UTF-8',
                        'no-outline': None,
                        'enable-local-file-access': None
                    }
                )
                return BytesIO(pdf_bytes)
            except Exception as pdf_error:
                logger.error(f"PDF yaratish xatosi (pdfkit): {pdf_error}")
                # pdfkit ishlamasa, fallbackga o'tamiz
                if fallback_lines is None:
                    fallback_lines = _html_to_plain_text_lines(html_content)
        
        # Agar html_content mavjud bo'lmasa yoki pdfkit ishlamasa - fallback
        if fallback_lines is None:
            if html_content:
                fallback_lines = _html_to_plain_text_lines(html_content)
            else:
                fallback_lines = ["Hisobotni yaratib bo'lmadi."]
        
        return _generate_pdf_with_reportlab(fallback_title, fallback_lines)
    
    except Exception as e:
        logger.error(f"PDF yaratish xatosi: {e}")
        return None


def _build_default_result_html(result_data):
    """Oddiy foydalanuvchi natijasi uchun HTML yaratish"""
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Test Natijasi</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 24px; }}
                h1 {{ color: #2c3e50; }}
                .info {{ background: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 6px; }}
                .question {{ margin: 15px 0; padding: 12px; background: #fff; border-left: 3px solid #007bff; border-radius: 4px; }}
                .correct {{ color: #2ecc71; }}
                .incorrect {{ color: #e74c3c; }}
            </style>
        </head>
        <body>
            <h1>Test Natijasi</h1>
            <div class="info">
                <p><strong>Test nomi:</strong> {result_data.get('test_name', "Noma'lum")}</p>
                <p><strong>Sana:</strong> {result_data.get('completed_at', '')}</p>
                <p><strong>To'g'ri javoblar:</strong> {result_data.get('correct', 0)}/{result_data.get('total', 0)}</p>
                <p><strong>Foiz:</strong> {result_data.get('percentage', 0):.1f}%</p>
            </div>
            <h2>Javoblar tafsiloti:</h2>
        """
        
        for idx, res in enumerate(result_data.get('results', []), 1):
            status = "✅ To'g'ri" if res.get('is_correct') else "❌ Noto'g'ri"
            status_class = "correct" if res.get('is_correct') else "incorrect"
            html_content += f"""
            <div class="question">
                <p><strong>Savol {idx}:</strong> {res.get('question', '')}</p>
                <p class="{status_class}"><strong>Javobingiz:</strong> {res.get('user_answer', '')} | <strong>To'g'ri javob:</strong> {res.get('correct_answer', '')} | {status}</p>
            </div>
            """
        
        html_content += """
        </body>
        </html>
        """
        return html_content
    except Exception as e:
        logger.error(f"HTML yaratish xatosi: {e}")
        return None


def _build_fallback_lines_from_result(result_data):
    """Oddiy foydalanuvchi natijasi uchun fallback matn"""
    lines = []
    try:
        test_name = result_data.get('test_name')
        if test_name:
            lines.append(f"Test nomi: {test_name}")
        completed_at = result_data.get('completed_at')
        if completed_at:
            lines.append(f"Sana: {completed_at}")
        correct = result_data.get('correct')
        total = result_data.get('total')
        if correct is not None and total is not None:
            lines.append(f"Natija: {correct}/{total} savol to'g'ri")
        percentage = result_data.get('percentage')
        if percentage is not None:
            lines.append(f"Foiz: {percentage:.1f}%")
        lines.append("")
        lines.append("Javoblar tafsiloti:")
        for idx, res in enumerate(result_data.get('results', []), 1):
            status = "To'g'ri" if res.get('is_correct') else "Noto'g'ri"
            lines.append(f"Savol {idx}: {res.get('question', '')}")
            lines.append(
                f"  Javobingiz: {res.get('user_answer', '-')}, To'g'ri javob: {res.get('correct_answer', '-')}, {status}"
            )
        return lines
    except Exception as e:
        logger.error(f"Fallback matn yaratish xatosi: {e}")
        return ["Natijalar mavjud emas."]


def _html_to_plain_text_lines(html_content):
    """HTML matnini oddiy matn satrlariga aylantirish"""
    try:
        if not html_content:
            return []
        # Taglarni yangi qator bilan almashtirish
        html = re.sub(r'<\s*(br|BR)\s*/?>', '\n', html_content)
        html = re.sub(r'</\s*(p|div|tr|h[1-6]|li|ul|ol|table|thead|tbody|tfoot)>', '\n', html)
        html = re.sub(r'<\s*li\s*>', '• ', html)
        text = re.sub(r'<[^>]+>', '', html)
        lines = [line.strip() for line in text.splitlines()]
        return [line for line in lines if line]
    except Exception as e:
        logger.error(f"HTML ni oddiy matnga o'tkazish xatosi: {e}")
        return []


def _generate_pdf_with_reportlab(title, lines):
    """Reportlab orqali oddiy PDF yaratish"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
    except Exception as import_error:
        logger.error(f"Reportlab kutubxonasini import qilib bo'lmadi: {import_error}")
        return None
    
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        x_margin = 20 * mm
        y_margin = 20 * mm
        y_position = height - y_margin
        
        # Sarlavha
        if title:
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x_margin, y_position, str(title))
            y_position -= 18
        
        c.setFont("Helvetica", 11)
        line_height = 13
        
        for raw_line in lines:
            wrapped_lines = textwrap.wrap(str(raw_line), width=90) or ['']
            for line in wrapped_lines:
                if y_position < y_margin:
                    c.showPage()
                    c.setFont("Helvetica", 11)
                    y_position = height - y_margin
                c.drawString(x_margin, y_position, line)
                y_position -= line_height
        
        c.save()
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Reportlab orqali PDF yaratish xatosi: {e}")
        return None


def generate_response_matrix(test_id, data):
    """0-1 matrix yaratish Excel formatida - ikkita alohida fayl
    
    Ikkita alohida Excel fayl yaratadi:
    1. matrix_1-40_*.xlsx - Questions 1-40 uchun
    2. matrix_41-43_*.xlsx - Questions 41-43 uchun
    
    Returns:
        tuple: (file_path_1_40, file_path_41_43, matrix_text) yoki (None, None, None)
    """
    try:
        # Test natijalarini to'plash
        test_results = [
            r for r_id, r in data.get('user_results', {}).items()
            if r.get('test_id') == test_id
        ]
        
        if not test_results:
            return None, None, None
        
        from openpyxl.styles import Font, Alignment
        
        matrix_dir = "matrices"
        os.makedirs(matrix_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        # ===== 1. Questions 1-40 uchun alohida Excel fayl =====
        wb_main = Workbook()
        
        # Standart sheetni o'chirish
        if 'Sheet' in wb_main.sheetnames:
            wb_main.remove(wb_main['Sheet'])
        
        ws_main = wb_main.create_sheet("Questions 1-40")
        
        # Header: Talabgor, Q1, Q2, ..., Q40
        main_header = ['Talabgor'] + [f'Q{i+1}' for i in range(40)]
        ws_main.append(main_header)
        
        # Header qatorini qalinlashtirish
        for cell in ws_main[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Har bir foydalanuvchi uchun qator (1-40 savollar)
        for result in test_results:
            user_id = result['user_id']
            # Foydalanuvchi ism-familyasini olish
            user_info = data.get('users', {}).get(str(user_id), {})
            first_name = user_info.get('first_name', '').strip()
            last_name = user_info.get('last_name', '').strip()
            full_name = f"{first_name} {last_name}".strip()
            # Agar ism-familya bo'lmasa, user_id ni ko'rsatish
            if not full_name:
                full_name = str(user_id)
            row = [full_name]
            
            # Birinchi 40 ta savol (0-39 indices)
            for i, res in enumerate(result['results'][:40]):
                value = 1 if res.get('is_correct') else 0
                row.append(value)
            
            ws_main.append(row)
        
        # Birinchi faylni saqlash
        file_path_1_40 = os.path.join(matrix_dir, f"matrix_1-40_{test_id}_{timestamp}.xlsx")
        wb_main.save(file_path_1_40)
        
        # ===== 2. Questions 41-43 uchun alohida Excel fayl =====
        # Har bir kichik savol uchun alohida ustun
        wb_problem = Workbook()
        
        # Standart sheetni o'chirish
        if 'Sheet' in wb_problem.sheetnames:
            wb_problem.remove(wb_problem['Sheet'])
        
        ws_problem = wb_problem.create_sheet("Questions 41-43")
        
        # Header yaratish: Talabgor, 41.1, 41.2, ..., 43.n
        problem_header = ['Talabgor']
        
        # Testdan masalaviy savollar strukturasini olish
        test = data.get('tests', {}).get(test_id)
        if test and test.get('questions'):
            problem_questions = [q for q in test['questions'][40:43] if q.get('type') == 'problem']
            
            for q_idx, question in enumerate(problem_questions, start=41):
                sub_count = question.get('sub_question_count', 0)
                if sub_count > 0:
                    # Har bir kichik savol uchun ustun
                    for sub_idx in range(1, sub_count + 1):
                        problem_header.append(f"{q_idx}.{sub_idx}")
                else:
                    # Eski format - faqat Q41, Q42, Q43
                    problem_header.append(f"Q{q_idx}")
        
        ws_problem.append(problem_header)
        
        # Header qatorini qalinlashtirish
        for cell in ws_problem[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Har bir foydalanuvchi uchun qator (41-43 savollar)
        for result in test_results:
            user_id = result['user_id']
            
            # Foydalanuvchi ism-familyasini olish
            user_info = data.get('users', {}).get(str(user_id), {})
            first_name = user_info.get('first_name', '').strip()
            last_name = user_info.get('last_name', '').strip()
            full_name = f"{first_name} {last_name}".strip()
            # Agar ism-familya bo'lmasa, user_id ni ko'rsatish
            if not full_name:
                full_name = str(user_id)
            row = [full_name]
            
            # 41-43 savollar uchun javoblar (indices 40-42)
            problem_results = result['results'][40:43] if len(result['results']) > 40 else []
            
            for q_idx, res in enumerate(problem_results, start=41):
                if res.get('type') == 'problem' and 'sub_results' in res:
                    # Yangi format - har bir kichik javob uchun alohida ustun
                    for sub in res['sub_results']:
                        value = 1 if sub.get('is_correct') else 0
                        row.append(value)
                else:
                    # Eski format - sub_results ni qayta yaratish
                    # user_answer va correct_answer ni taqqoslash
                    if test and len(test.get('questions', [])) > 40:
                        original_q_idx = 40 + (q_idx - 41)
                        question = test['questions'][original_q_idx]
                        
                        # Testdan to'g'ri javoblar ro'yxatini olish
                        correct_answers = question.get('correct', [])
                        
                        # Foydalanuvchi javoblarini olish
                        user_answer = res.get('user_answer', '')
                        user_answers_list = [a.strip() for a in user_answer.split(',')] if user_answer else []
                        
                        # Har bir javobni taqqoslash
                        for i, correct_ans in enumerate(correct_answers):
                            if i < len(user_answers_list):
                                user_ans = user_answers_list[i]
                                is_sub_correct = user_ans.lower() == correct_ans.strip().lower()
                                row.append(1 if is_sub_correct else 0)
                            else:
                                # Javob berilmagan
                                row.append(0)
                    else:
                        # Test topilmasa, faqat umumiy natijani qo'shamiz
                        value = 1 if res.get('is_correct') else 0
                        row.append(value)
            
            ws_problem.append(row)
        
        # Ikkinchi faylni saqlash
        file_path_41_43 = os.path.join(matrix_dir, f"matrix_41-43_{test_id}_{timestamp}.xlsx")
        wb_problem.save(file_path_41_43)
        
        # Matn formatini ham saqlash (1-40 savollar uchun)
        matrix_lines = []
        matrix_lines.append('\t'.join(main_header))
        for result in test_results:
            user_id = result['user_id']
            # Foydalanuvchi ism-familyasini olish
            user_info = data.get('users', {}).get(str(user_id), {})
            first_name = user_info.get('first_name', '').strip()
            last_name = user_info.get('last_name', '').strip()
            full_name = f"{first_name} {last_name}".strip()
            # Agar ism-familya bo'lmasa, user_id ni ko'rsatish
            if not full_name:
                full_name = str(user_id)
            row = [full_name]
            for i, res in enumerate(result['results'][:40]):
                value = '1' if res.get('is_correct') else '0'
                row.append(value)
            matrix_lines.append('\t'.join(row))
        matrix_text = '\n'.join(matrix_lines)
        
        return file_path_1_40, file_path_41_43, matrix_text
        
    except Exception as e:
        logger.error(f"Matrix yaratish xatosi: {e}")
        return None, None, None


def rasch_model_analysis(data_matrix):
    """
    Rasch model (1PL IRT) tahlili
    
    Parameters:
    - data_matrix: Numpy array (qatorlar: talabalar, ustunlar: savollar), 0/1
    
    Returns:
    - theta: Talabalar qobiliyati (float32)
    - beta: Savollar qiyinligi (float32)
    """
    n_students, n_items = data_matrix.shape
    
    # Boshlang'ich baholar
    student_scores = np.sum(data_matrix, axis=1, dtype=np.float64)
    item_scores = np.sum(data_matrix, axis=0, dtype=np.float64)
    
    # Theta (talaba qobiliyatlari) - logit transformatsiya
    theta = np.zeros(n_students, dtype=np.float64)
    for i in range(n_students):
        if student_scores[i] == 0:
            theta[i] = -3.0
        elif student_scores[i] == n_items:
            theta[i] = 3.0
        else:
            p = (student_scores[i] + 0.5) / (n_items + 1)
            p = np.clip(p, 1e-6, 1 - 1e-6)
            theta[i] = np.log(p / (1 - p))
    
    # Beta (savol qiyinliklari) - logit transformatsiya
    beta = np.zeros(n_items, dtype=np.float64)
    for j in range(n_items):
        if item_scores[j] == 0:
            beta[j] = 3.0
        elif item_scores[j] == n_students:
            beta[j] = -3.0
        else:
            p = (item_scores[j] + 0.5) / (n_students + 1)
            p = np.clip(p, 1e-6, 1 - 1e-6)
            beta[j] = -np.log(p / (1 - p))
    
    # MLE iteratsiyalari (Rasch model uchun)
    max_iter = 100
    tol = 1e-6
    REG_LAMBDA = 0.05
    
    for iteration in range(max_iter):
        old_theta = theta.copy()
        old_beta = beta.copy()
        
        # Ehtimolliklar hisoblash
        logits = theta[:, np.newaxis] - beta[np.newaxis, :]
        np.clip(logits, -15, 15, out=logits)
        p = expit(logits)
        residuals = data_matrix - p
        
        # Theta yangilanishi
        grad_theta = np.sum(residuals, axis=1) - REG_LAMBDA * theta
        hess_theta = np.sum(p * (1 - p), axis=1) + REG_LAMBDA
        update_theta = np.where(hess_theta > 1e-10, grad_theta / hess_theta, 0.0)
        theta += update_theta
        
        # Beta yangilanishi
        grad_beta = -np.sum(residuals, axis=0) - REG_LAMBDA * beta
        hess_beta = np.sum(p * (1 - p), axis=0) + REG_LAMBDA
        update_beta = np.where(hess_beta > 1e-10, grad_beta / hess_beta, 0.0)
        beta += update_beta
        
        # Konvergensiya tekshiruvi
        if max(np.max(np.abs(update_theta)), np.max(np.abs(update_beta))) < tol:
            break
    
    # Identifikatsiya: theta ni markazlash (mean = 0)
    theta = theta - np.mean(theta)
    
    return theta.astype(np.float32), beta.astype(np.float32)


def ability_to_standard_score(ability):
    """
    Qobiliyatni standart ballga o'tkazish (T-score)
    
    Parameters:
    - ability: Talabaning qobiliyat bahosi (θ)
    
    Returns:
    - standard_score: Standart ball (0-100)
    """
    t_score = 50 + (10 * ability)
    standard_score = np.clip(t_score, 0, 100)
    return standard_score


def ability_to_grade(ability):
    """
    Qobiliyatni bahoga o'tkazish
    
    Parameters:
    - ability: Talabaning qobiliyat bahosi (θ)
    
    Returns:
    - grade: Tayinlangan baho
    """
    t_score = ability_to_standard_score(ability)
    
    if isinstance(t_score, np.ndarray):
        grades = np.where(t_score >= 70, 'A+',
                 np.where(t_score >= 65, 'A',
                 np.where(t_score >= 60, 'B+',
                 np.where(t_score >= 55, 'B',
                 np.where(t_score >= 50, 'C+',
                 np.where(t_score >= 46, 'C', 'NC'))))))
        return grades
    else:
        if t_score >= 70:
            return 'A+'
        elif t_score >= 65:
            return 'A'
        elif t_score >= 60:
            return 'B+'
        elif t_score >= 55:
            return 'B'
        elif t_score >= 50:
            return 'C+'
        elif t_score >= 46:
            return 'C'
        else:
            return 'NC'


def perform_rasch_analysis(test_id, data, question_range='1-40'):
    """
    Test natijalarini Rasch modelida tahlil qilish
    
    Parameters:
    - test_id: Test ID
    - data: Ma'lumotlar bazasi
    - question_range: '1-40' yoki '40-43' (yozma savollar)
    
    Returns:
    - dict: Rasch tahlil natijalari
    """
    try:
        # Test natijalarini to'plash
        test_results = [
            r for r_id, r in data.get('user_results', {}).items()
            if r.get('test_id') == test_id
        ]
        
        if not test_results or len(test_results) < 2:
            return None  # Kamida 2 ta natija kerak
        
        test = data.get('tests', {}).get(test_id)
        if not test:
            return None
        
        user_ids = []
        
        if question_range == '1-40':
            # 1-40 savollar uchun (ko'p tanlov)
            n_students = len(test_results)
            n_items = 40
            
            data_matrix = np.zeros((n_students, n_items), dtype=np.int32)
            
            for idx, result in enumerate(test_results):
                user_ids.append(result['user_id'])
                # Birinchi 40 ta savol uchun javoblar
                for i, res in enumerate(result['results'][:40]):
                    data_matrix[idx, i] = 1 if res.get('is_correct') else 0
            
            # Rasch model tahlili
            theta, beta = rasch_model_analysis(data_matrix)
            
            # Har bir talaba uchun standart ball va baho
            standard_scores = ability_to_standard_score(theta)
            grades = ability_to_grade(theta)
            
            return {
                'user_ids': user_ids,
                'abilities': theta.tolist(),
                'standard_scores': standard_scores.tolist() if isinstance(standard_scores, np.ndarray) else [standard_scores],
                'grades': grades.tolist() if isinstance(grades, np.ndarray) else [grades],
                'item_difficulties': beta.tolist(),
                'n_students': n_students,
                'n_items': n_items
            }
        
        elif question_range == '40-43':
            # 40-43 savollar uchun (yozma savollar) - Rasch modelida
            n_students = len(test_results)
            
            # 40-43 savollar sonini aniqlash va 0-1 matrix yaratish
            text_questions = [q for q in test.get('questions', [])[35:40] if q.get('type') == 'text_answer']
            problem_questions = [q for q in test.get('questions', [])[40:43] if q.get('type') == 'problem']
            
            # Har bir masalaviy savol uchun kichik savollar sonini hisoblash
            total_sub_items = 0
            for q in problem_questions:
                total_sub_items += q.get('sub_question_count', 1)
            
            n_text_items = len(text_questions)
            total_items = n_text_items + total_sub_items
            
            if total_items == 0:
                return None
            
            # 0-1 matrix yaratish (har bir kichik savol uchun alohida ustun)
            data_matrix = np.zeros((n_students, total_items), dtype=np.int32)
            
            # Avval barcha user_ids ni to'plash
            for result in test_results:
                user_ids.append(result['user_id'])
            
            # Keyin matrix yaratish
            for idx, result in enumerate(test_results):
                item_idx = 0
                
                # 36-40 yozma savollar (indices 35-39)
                for i in range(35, min(40, len(result['results']))):
                    res = result['results'][i]
                    if res.get('is_correct'):
                        data_matrix[idx, item_idx] = 1
                    item_idx += 1
                
                # 41-43 masalaviy savollar (indices 40-42)
                for i in range(40, min(43, len(result['results']))):
                    res = result['results'][i]
                    if res.get('type') == 'problem' and 'sub_results' in res:
                        # Har bir kichik savol uchun alohida ustun
                        for sub in res['sub_results']:
                            if sub.get('is_correct'):
                                data_matrix[idx, item_idx] = 1
                            item_idx += 1
                    else:
                        # Oddiy savol
                        if res.get('is_correct'):
                            data_matrix[idx, item_idx] = 1
                        item_idx += 1
            
            # Rasch model tahlili
            theta, beta = rasch_model_analysis(data_matrix)
            
            # Har bir talaba uchun standart ball (0-100)
            standard_scores = ability_to_standard_score(theta)
            
            # 0-100 dan 0-75 ga o'tkazish
            written_scores_scaled = [score * 0.75 for score in (standard_scores.tolist() if isinstance(standard_scores, np.ndarray) else [standard_scores])]
            
            return {
                'user_ids': user_ids,
                'abilities': theta.tolist(),
                'written_scores': written_scores_scaled,  # 0-75 shkalada (Rasch model asosida)
                'written_scores_raw': standard_scores.tolist() if isinstance(standard_scores, np.ndarray) else [standard_scores],  # 0-100 shkalada (Rasch model)
                'item_difficulties': beta.tolist(),
                'n_students': n_students,
                'n_items': total_items
            }
        
        return None
        
    except Exception as e:
        logger.error(f"Rasch tahlil xatosi: {e}")
        return None


def generate_final_results_excel(test_id, data):
    """
    Yakuniy natijalar Excel faylini yaratish
    
    Format:
    Talabgor | Test-ball | Yozma ball | Yakuniy ball | Daraja | Foiz
    
    Parameters:
    - test_id: Test ID
    - data: Ma'lumotlar bazasi
    
    Returns:
    - str: Excel fayl yo'li yoki None
    """
    try:
        # 1-40 savollar uchun Rasch tahlili
        test_results_1_40 = perform_rasch_analysis(test_id, data, '1-40')
        # 40-43 savollar uchun yozma ball
        test_results_40_43 = perform_rasch_analysis(test_id, data, '40-43')
        
        if not test_results_1_40:
            return None
        
        # Barcha foydalanuvchilar ro'yxatini olish
        all_user_ids = test_results_1_40['user_ids']
        
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Yakuniy Natijalar"
        
        # Header
        headers = ['Talabgor', 'Test-ball', 'Yozma ball', 'Yakuniy ball', 'Daraja', 'Foiz']
        ws.append(headers)
        
        # Header qatorini formatlash
        from openpyxl.styles import Font, Alignment, PatternFill
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
        
        # User ID bo'yicha mapping
        test_scores_map = {}
        standard_scores_list = test_results_1_40['standard_scores']
        if not isinstance(standard_scores_list, list):
            standard_scores_list = [standard_scores_list]
        
        for i, user_id in enumerate(test_results_1_40['user_ids']):
            test_scores_map[user_id] = standard_scores_list[i] if i < len(standard_scores_list) else 0
        
        written_scores_map = {}
        if test_results_40_43:
            written_scores_list = test_results_40_43['written_scores']
            if not isinstance(written_scores_list, list):
                written_scores_list = [written_scores_list]
            
            for i, user_id in enumerate(test_results_40_43['user_ids']):
                written_scores_map[user_id] = written_scores_list[i] if i < len(written_scores_list) else 0
        
        # Har bir foydalanuvchi uchun qator
        for user_id in all_user_ids:
            # Foydalanuvchi ism-familiyasi
            user_info = data.get('users', {}).get(str(user_id), {})
            first_name = user_info.get('first_name', '').strip()
            last_name = user_info.get('last_name', '').strip()
            full_name = f"{first_name} {last_name}".strip() if first_name and last_name else f"User {user_id}"
            
            # Test-ball (1-40 dan, Rasch model)
            test_score = test_scores_map.get(user_id, 0)
            if isinstance(test_score, dict):
                test_score = test_score.get('test_score', 0)
            
            # Yozma ball (40-43 dan, 0-75 shkalada)
            written_score = written_scores_map.get(user_id, 0)
            
            # Yakuniy ball (ikkalasining o'rtachasi)
            final_score = (test_score + written_score) / 2
            
            # Daraja (yakuniy ball asosida)
            final_grade = ability_to_grade_from_score(final_score)
            
            # Foiz (yakuniy_ball / 65 * 100)
            percentage = (final_score / 65) * 100 if 65 > 0 else 0
            percentage = min(percentage, 100)  # 100% dan oshmasligi kerak
            
            # Qator qo'shish
            row = [
                full_name,
                round(test_score, 2),
                round(written_score, 2),
                round(final_score, 2),
                final_grade,
                round(percentage, 2)
            ]
            ws.append(row)
        
        # Ustunlarni kengaytirish
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20
        
        # Faylni saqlash
        results_dir = "final_results"
        os.makedirs(results_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        file_path = os.path.join(results_dir, f"final_results_{test_id}_{timestamp}.xlsx")
        wb.save(file_path)
        
        return file_path
        
    except Exception as e:
        logger.error(f"Yakuniy natijalar Excel yaratish xatosi: {e}")
        return None


def ability_to_grade_from_score(score):
    """
    Ball asosida baho aniqlash
    
    Parameters:
    - score: Ball (0-100)
    
    Returns:
    - grade: Baho
    """
    if score >= 70:
        return 'A+'
    elif score >= 65:
        return 'A'
    elif score >= 60:
        return 'B+'
    elif score >= 55:
        return 'B'
    elif score >= 50:
        return 'C+'
    elif score >= 46:
        return 'C'
    else:
        return 'NC'


