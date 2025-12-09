#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bot handler funksiyalari
"""

import logging
import re
import os
from datetime import datetime, timedelta
import pytz
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import ContextTypes
import pdfkit

from config import BOSS_ID
from database import load_data, save_data
from utils import check_subscription, generate_pdf

# O'zbekiston vaqti (UTC+5)
UZBEKISTAN_TZ = pytz.timezone('Asia/Tashkent')

logger = logging.getLogger(__name__)


def generate_test_post(test_data, test_id, bot_username=None):
    """
    Test haqida ko'rkam va tushunarli post yaratish

    Args:
        test_data: Test ma'lumotlari
        test_id: Test ID
        bot_username: Bot username (ixtiyoriy)

    Returns:
        tuple: (post_text, inline_keyboard) - HTML formatidagi post matni va inline keyboard
    """
    test_name = test_data.get('name', 'Test')
    total_questions = len(test_data.get('questions', []))

    # Bot username (agar berilmagan bo'lsa, default)
    if not bot_username:
        bot_username = "Test Bot"

    post = f"""
📝 <b>{test_name}</b>

📊 <b>Ma'lumotlar:</b>
• Savollar soni: {total_questions} ta
• Format: Ko'p tanlovli (a, b, c, d)

⚠️ <b>Shartlar:</b>
• Har bir testni faqat <b>1 marta</b> ishlash mumkin
• Javoblarni <b>1a2b3c4d...</b> formatida yuboring
• Javoblar soni savollar soniga mos kelishi kerak

✅ Test yakunlangach, natijalar o'qituvchi tomonidan e'lon qilinadi.

🎓 Muvaffaqiyatlar!
"""

    # Inline keyboard yaratish - testni boshlash uchun
    keyboard = [
        [InlineKeyboardButton("🚀 Testni boshlash", callback_data=f"start_test_{test_id}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    return post, reply_markup


async def check_user_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Foydalanuvchining ism va familyasini tekshirish"""
    user_id = update.effective_user.id
    data = load_data()

    # Foydalanuvchi ma'lumotlarini tekshirish
    user_info = data.get('users', {}).get(str(user_id), {})
    first_name = user_info.get('first_name', '').strip()
    last_name = user_info.get('last_name', '').strip()

    # Ism va familya to'liq bo'lishi kerak
    if not first_name or not last_name:
        return False

    # Ism va familya kamida 2 ta harfdan iborat bo'lishi kerak
    if len(first_name) < 2 or len(last_name) < 2:
        return False

    return True


async def process_user_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ism va familya kiritish jarayonini qayta ishlash"""
    if not update.message or not update.message.text:
        return

    text = update.message.text.strip()
    user_id = update.effective_user.id
    step = context.user_data.get('name_step', 'first_name')

    # Ism kiritish
    if step == 'first_name':
        # Ism validatsiyasi
        if len(text) < 2:
            await update.message.reply_text(
                "❌ Ism kamida 2 ta harfdan iborat bo'lishi kerak.\n\n"
                "📝 Ismingizni qayta kiriting:"
            )
            return

        # Raqamlar yoki maxsus belgilar tekshiruvi (faqat harflar va probellar)
        if not all(c.isalpha() or c.isspace() for c in text):
            await update.message.reply_text(
                "❌ Ism faqat harflardan iborat bo'lishi kerak.\n\n"
                "📝 Ismingizni qayta kiriting:"
            )
            return

        # Ismni saqlash
        context.user_data['first_name'] = text.strip()
        context.user_data['name_step'] = 'last_name'

        await update.message.reply_text(
            "✅ Ism qabul qilindi!\n\n"
            "📝 Familyangizni kiriting:"
        )
        return

    # Familya kiritish
    elif step == 'last_name':
        # Familya validatsiyasi
        if len(text) < 2:
            await update.message.reply_text(
                "❌ Familya kamida 2 ta harfdan iborat bo'lishi kerak.\n\n"
                "📝 Familyangizni qayta kiriting:"
            )
            return

        # Raqamlar yoki maxsus belgilar tekshiruvi (faqat harflar va probellar)
        if not all(c.isalpha() or c.isspace() for c in text):
            await update.message.reply_text(
                "❌ Familya faqat harflardan iborat bo'lishi kerak.\n\n"
                "📝 Familyangizni qayta kiriting:"
            )
            return

        # Familyani saqlash
        first_name = context.user_data.get('first_name', '').strip()
        last_name = text.strip()

        # Ma'lumotlar bazasiga saqlash
        data = load_data()
        if 'users' not in data:
            data['users'] = {}

        data['users'][str(user_id)] = {
            'first_name': first_name,
            'last_name': last_name,
            'registered_at': datetime.now().isoformat()
        }
        save_data(data)

        # User data ni tozalash
        context.user_data.pop('waiting_for_name', None)
        context.user_data.pop('name_step', None)
        context.user_data.pop('first_name', None)

        # Xush kelibsiz xabari
        full_name = f"{first_name} {last_name}"
        keyboard = [
            [KeyboardButton("📝 Test ishlash"), KeyboardButton("📊 Test natijalarim")]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

        await update.message.reply_text(
            f"✅ {full_name}, ma'lumotlaringiz qabul qilindi!\n\n"
            f"Endi botdan foydalanishingiz mumkin.",
            reply_markup=reply_markup
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start komandasi"""
    user_id = update.effective_user.id

    # Majburiy obuna tekshiruvi
    if not await check_subscription(update, context):
        data = load_data()

        # Inline keyboard yaratish kanallar uchun
        keyboard = []
        for ch in data["mandatory_channels"]:
            # Kanal username yoki ID ni tayyorlash
            if ch.startswith('@'):
                channel_username = ch.replace('@', '')
                button_text = ch
                # Inline button yaratish (kanalga havola)
                keyboard.append([InlineKeyboardButton(
                    text=f"📢 {button_text}",
                    url=f"https://t.me/{channel_username}"
                )])
            elif ch.startswith('-'):
                # Kanal ID bo'lsa, button yaratib bo'lmaydi (URL kerak emas)
                # Faqat matn sifatida ko'rsatamiz
                pass
            else:
                # Username @ belgisiz bo'lsa
                channel_username = ch
                button_text = f"@{ch}"
                # Inline button yaratish (kanalga havola)
                keyboard.append([InlineKeyboardButton(
                    text=f"📢 {button_text}",
                    url=f"https://t.me/{channel_username}"
                )])

        # Inline keyboard markup yaratish
        reply_markup = InlineKeyboardMarkup(keyboard) if keyboard else None

        # Kanal ro'yxatini matn sifatida ham ko'rsatish
        channels_text = "\n".join([
            f"• {ch if ch.startswith('@') else f'@{ch}' if not ch.startswith('-') else f'Kanal ID: {ch}'}"
            for ch in data["mandatory_channels"]
        ])

        text = f"❌ Botdan foydalanish uchun quyidagi kanallarga obuna bo'lishingiz kerak:\n\n{channels_text}\n\n"
        text += "Quyidagi tugmalar orqali kanallarga o'ting va obuna bo'ling.\n"
        text += "Obuna bo'lgach, /start ni qayta bosing."

        await update.message.reply_text(text, reply_markup=reply_markup)
        return

    # Ism va familya tekshiruvi
    if not await check_user_name(update, context):
        # Ism va familya kiritish rejimini boshlash
        context.user_data['waiting_for_name'] = True
        context.user_data['name_step'] = 'first_name'

        await update.message.reply_text(
            "👤 Botdan foydalanish uchun ism va familyangizni kiriting.\n\n"
            "Iltimos, to'g'ri ism va familyangizni kiriting:\n\n"
            "📝 Ismingizni kiriting:"
        )
        return

    # Foydalanuvchi turini aniqlash
    data = load_data()
    is_boss = user_id == BOSS_ID
    is_admin = user_id in data["admins"]

    # Reply keyboard markup yaratish
    keyboard = [
        [KeyboardButton("📝 Test ishlash"), KeyboardButton("📊 Test natijalarim")]
    ]
    # Faqat adminlar va boss uchun qo'shimcha tugmalar
    if is_boss or is_admin:
        keyboard.append([KeyboardButton("➕ Test yaratish"), KeyboardButton("📈 Statistika")])
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    # Foydalanuvchi ismi va familyasini olish
    user_info = data.get('users', {}).get(str(user_id), {})
    first_name = user_info.get('first_name', '')
    last_name = user_info.get('last_name', '')
    full_name = f"{first_name} {last_name}".strip()

    if is_boss:
        text = f"👑 {full_name}, Boss paneliga xush kelibsiz!\n\n"
        text += "/admin - Adminlar boshqaruvi\n"
        text += "/channels - Kanal boshqaruvi\n"
        text += "/createtest - Test yaratish\n\n"
        text += "Yoki quyidagi tugmalardan foydalaning:"
    elif is_admin:
        text = f"👨‍💼 {full_name}, Admin paneliga xush kelibsiz!\n\n"
        text += "/createtest - Test yaratish\n\n"
        text += "Yoki quyidagi tugmalardan foydalaning:"
    else:
        text = f"👋 {full_name}, Test botiga xush kelibsiz!\n\n"
        text += "Quyidagi tugmalardan foydalaning:"

    await update.message.reply_text(text, reply_markup=reply_markup)


async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin panel (faqat boss uchun)"""
    if update.effective_user.id != BOSS_ID:
        await update.message.reply_text("❌ Bu funksiya faqat boss uchun!")
        return

    data = load_data()
    keyboard = [
        [InlineKeyboardButton("➕ Admin qo'shish", callback_data="add_admin")],
        [InlineKeyboardButton("➖ Admin olib tashlash", callback_data="remove_admin")],
        [InlineKeyboardButton("📋 Adminlar ro'yxati", callback_data="list_admins")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("👑 Admin boshqaruvi:", reply_markup=reply_markup)


async def channels_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kanal boshqaruvi (faqat boss uchun)"""
    if update.effective_user.id != BOSS_ID:
        await update.message.reply_text("❌ Bu funksiya faqat boss uchun!")
        return

    data = load_data()
    keyboard = [
        [InlineKeyboardButton("➕ Kanal qo'shish", callback_data="add_channel")],
        [InlineKeyboardButton("➖ Kanal olib tashlash", callback_data="remove_channel")],
        [InlineKeyboardButton("📋 Kanallar ro'yxati", callback_data="list_channels")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("📢 Kanal boshqaruvi:", reply_markup=reply_markup)


async def create_test(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test yaratish"""
    user_id = update.effective_user.id
    data = load_data()

    if user_id != BOSS_ID and user_id not in data["admins"]:
        await update.message.reply_text("❌ Bu funksiya faqat adminlar uchun!")
        return

    # Test yaratish rejimini boshlash
    context.user_data['creating_test'] = True
    context.user_data['test_creation_step'] = 'name'

    await update.message.reply_text(
        "📝 Test yaratish:\n\n"
        "Test nomini kiriting:\n\n"
        "Yoki /cancel ni bosing bekor qilish uchun."
    )


async def process_test_editing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test tahrirlash jarayonini qayta ishlash"""
    if not context.user_data.get('editing_test'):
        return

    if not update.message or not update.message.text:
        return

    user_id = update.effective_user.id
    test_id = context.user_data.get('editing_test_id')
    step = context.user_data.get('test_editing_step', 'name')

    data = load_data()
    if test_id not in data['tests']:
        await update.message.reply_text("❌ Test topilmadi!")
        context.user_data.pop('editing_test', None)
        return

    test = data['tests'][test_id]

    if step == 'name':
        # Test nomi o'zgartirildi
        new_name = update.message.text.strip()
        if new_name == '/cancel':
            context.user_data.pop('editing_test', None)
            context.user_data.pop('editing_test_id', None)
            context.user_data.pop('test_editing_step', None)
            await update.message.reply_text("❌ Tahrirlash bekor qilindi.")
            return

        test['name'] = new_name
        save_data(data)
        context.user_data.pop('editing_test', None)
        context.user_data.pop('editing_test_id', None)
        context.user_data.pop('test_editing_step', None)
        await update.message.reply_text(f"✅ Test nomi o'zgartirildi: {new_name}")
        return

    elif step == 'answers':
        # Javoblar o'zgartirildi
        answers_text = update.message.text.strip().lower()
        if answers_text == '/cancel':
            context.user_data.pop('editing_test', None)
            context.user_data.pop('editing_test_id', None)
            context.user_data.pop('test_editing_step', None)
            await update.message.reply_text("❌ Tahrirlash bekor qilindi.")
            return

        try:
            questions = test['questions']
            answers = []

            for char in answers_text:
                if char.lower() in 'abcd':
                    answers.append(char.lower())

            if len(answers) != len(questions):
                await update.message.reply_text(
                    f"❌ Javoblar soni noto'g'ri!\n"
                    f"Savollar soni: {len(questions)}\n"
                    f"Javoblar soni: {len(answers)}\n\n"
                    f"Qayta kiriting: 1a2b3c4d... formatida"
                )
                return

            # Javoblarni yangilash
            for idx, answer in enumerate(answers):
                if idx < len(questions):
                    questions[idx]['correct'] = answer

            test['questions'] = questions
            save_data(data)
            context.user_data.pop('editing_test', None)
            context.user_data.pop('editing_test_id', None)
            context.user_data.pop('test_editing_step', None)
            await update.message.reply_text(f"✅ Javoblar yangilandi!")

        except Exception as e:
            logger.error(f"Javoblar yangilash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")




async def save_test_immediately(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Testni darhol saqlash (vaqt belgilamasdan)"""
    user_id = update.effective_user.id

    # Testni saqlash
    data = load_data()
    test_id = f"test_{len(data['tests']) + 1}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    test_data = {
        'name': context.user_data['test_name'],
        'questions': context.user_data['test_questions'],
        'created_by': user_id,
        'created_at': datetime.now(UZBEKISTAN_TZ).isoformat()
    }
    # Agar fayl yuklangan bo'lsa, faylni test ID bilan qayta nomlash
    if 'test_file_path' in context.user_data:
        old_file_path = context.user_data['test_file_path']
        test_files_dir = "test_files"
        file_name = context.user_data.get('test_file_name', 'test.txt')
        # Fayl kengaytmasini saqlash
        file_ext = os.path.splitext(file_name)[1] or '.txt'
        new_file_path = os.path.join(test_files_dir, f"{test_id}{file_ext}")

        # Faylni qayta nomlash
        if os.path.exists(old_file_path):
            os.rename(old_file_path, new_file_path)
            test_data['file_path'] = new_file_path
        else:
            test_data['file_path'] = old_file_path
        test_data['file_name'] = file_name
    data['tests'][test_id] = test_data
    save_data(data)

    test_name = context.user_data.get('test_name', 'Noma\'lum')
    test_questions = context.user_data.get('test_questions', [])

    # Test post yaratish va yuborish
    # Bot username ni olish
    try:
        bot_info = await context.bot.get_me()
        bot_username = f"@{bot_info.username}" if bot_info.username else "Test Bot"
    except:
        bot_username = "Test Bot"

    post_text, reply_markup = generate_test_post(test_data, test_id, bot_username)

    # Xabar yuborish
    await update.message.reply_text(
        post_text,
        parse_mode='HTML',
        reply_markup=reply_markup
    )

    # Test yaratildi xabari
    await update.message.reply_text(
        f"✅ Test muvaffaqiyatli yaratildi!\n\n"
        f"Test ID: {test_id}\n"
        f"Test nomi: {test_name}\n"
        f"Savollar soni: {len(test_questions)}"
    )

    context.user_data.clear()


async def process_test_creation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test yaratish jarayonini qayta ishlash"""
    if not context.user_data.get('creating_test'):
        return

    if not update.message or not update.message.text:
        return

    user_id = update.effective_user.id
    step = context.user_data.get('test_creation_step', 'name')

    if step == 'name':
        # Test nomi kiritildi
        test_name = update.message.text.strip()
        if test_name == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        # Test nomini saqlash va keyingi bosqichga o'tish (fayl so'ralmaydi)
        context.user_data['test_name'] = test_name
        context.user_data['test_creation_step'] = 'answers'
        await update.message.reply_text(
            "✅ Test nomi qabul qilindi!\n\n"
            "📝 Endi 1-35 savollar uchun javoblarni kiriting:\n\n"
            "Format: 1a2b3c4d... yoki abc... (35 ta javob)\n"
            "⚠️ 33, 34, 35-savollar uchun e va f javoblar ham mumkin!"
        )
        return

    elif step == 'text_answers':
        # 36-40 savollar uchun yozma javoblar kiritildi
        text_answers_input = update.message.text
        if text_answers_input.strip() == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        try:
            # Yozma javoblarni qatorlarga ajratish
            # Bo'sh qatorlar ham qabul qilinadi (5 ta qator bo'lishi kerak)
            lines = text_answers_input.split('\n')

            # Faqat birinchi 5 ta qatorni olish
            text_answers = []
            for i in range(min(5, len(lines))):
                line = lines[i].strip()
                text_answers.append(line)

            # Agar 5 ta qatordan kam bo'lsa, xatolik berish
            if len(text_answers) < 5:
                await update.message.reply_text(
                    f"❌ Yozma javoblar soni noto'g'ri!\n\n"
                    f"36-40 savollar uchun 5 ta qator kerak.\n"
                    f"Hozirgi son: {len(text_answers)} ta qator\n\n"
                    f"Har bir savol uchun alohida qatorda javob yozing:\n\n"
                    f"Masalan:\n"
                    f"ahsb\n"
                    f"hhhsbb\n"
                    f"\n"
                    f"uuus77\n"
                    f"\n\n"
                    f"ℹ️ Bo'sh qatorlar javob berilmagan degan ma'noni bildiradi."
                )
                return

            # Bo'sh javoblarni tekshirish - barcha javoblar majburiy
            empty_questions = []
            for i, answer in enumerate(text_answers):
                if not answer:  # Bo'sh qator
                    empty_questions.append(36 + i)

            if empty_questions:
                await update.message.reply_text(
                    f"❌ Barcha savollarga javob berish majburiy!\n\n"
                    f"Quyidagi savollar uchun javoblar kiritilmagan:\n"
                    f"{', '.join(map(str, empty_questions))}-savollar\n\n"
                    f"Iltimos, barcha 5 ta savol uchun javoblarni kiriting."
                )
                return

            # Yozma javoblarni saqlash
            context.user_data['text_answers'] = text_answers

            # 36-40 javoblarni saqlash va 41-savol uchun javoblarni so'rash
            context.user_data['test_creation_step'] = 'problem_41_answers'
            await update.message.reply_text(
                f"✅ 36-40 savollar uchun javoblar qabul qilindi!\n\n"
                f"📝 Endi 41-savol (masalaviy savol) uchun javoblarni kiriting:\n\n"
                f"⚠️ Har bir kichik savol uchun alohida qatorda javob yozing:\n\n"
                f"Masalan:\n"
                f"147782\n"
                f"H20\n"
                f"H2o\n"
                f"Na2\n"
                f"Ejeh\n\n"
                f"⚠️ Barcha javoblarni majburiy ravishda kiriting!"
            )
            return

        except Exception as e:
            logger.error(f"Yozma javoblar qayta ishlash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")

    elif step == 'problem_41_answers':
        # 41-savol (masalaviy) uchun javoblar kiritildi
        problem_41_input = update.message.text
        if problem_41_input.strip() == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        try:
            # Javoblarni qatorlarga ajratish
            lines = problem_41_input.split('\n')
            problem_41_answers = []

            for line in lines:
                line = line.strip()
                if line:  # Faqat bo'sh bo'lmagan qatorlarni olish
                    problem_41_answers.append(line)

            # Hech bo'lmaganda 1 ta javob bo'lishi kerak
            if len(problem_41_answers) == 0:
                await update.message.reply_text(
                    f"❌ 41-savol uchun hech bo'lmaganda 1 ta javob kiritish kerak!\n\n"
                    f"Qaytadan urinib ko'ring."
                )
                return

            # 41-savol javoblarini saqlash
            context.user_data['problem_41_answers'] = problem_41_answers

            # 42-savol uchun javoblarni so'rash
            context.user_data['test_creation_step'] = 'problem_42_answers'
            await update.message.reply_text(
                f"✅ 41-savol uchun {len(problem_41_answers)} ta javob qabul qilindi!\n\n"
                f"📝 Endi 42-savol (masalaviy savol) uchun javoblarni kiriting:\n\n"
                f"⚠️ Har bir kichik savol uchun alohida qatorda javob yozing:\n\n"
                f"Masalan:\n"
                f"javob1\n"
                f"javob2\n"
                f"javob3\n\n"
                f"⚠️ Barcha javoblarni majburiy ravishda kiriting!"
            )
            return

        except Exception as e:
            logger.error(f"41-savol javoblarini qayta ishlash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")

    elif step == 'problem_42_answers':
        # 42-savol (masalaviy) uchun javoblar kiritildi
        problem_42_input = update.message.text
        if problem_42_input.strip() == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        try:
            # Javoblarni qatorlarga ajratish
            lines = problem_42_input.split('\n')
            problem_42_answers = []

            for line in lines:
                line = line.strip()
                if line:  # Faqat bo'sh bo'lmagan qatorlarni olish
                    problem_42_answers.append(line)

            # Hech bo'lmaganda 1 ta javob bo'lishi kerak
            if len(problem_42_answers) == 0:
                await update.message.reply_text(
                    f"❌ 42-savol uchun hech bo'lmaganda 1 ta javob kiritish kerak!\n\n"
                    f"Qaytadan urinib ko'ring."
                )
                return

            # 42-savol javoblarini saqlash
            context.user_data['problem_42_answers'] = problem_42_answers

            # 43-savol uchun javoblarni so'rash
            context.user_data['test_creation_step'] = 'problem_43_answers'
            await update.message.reply_text(
                f"✅ 42-savol uchun {len(problem_42_answers)} ta javob qabul qilindi!\n\n"
                f"📝 Endi 43-savol (masalaviy savol) uchun javoblarni kiriting:\n\n"
                f"⚠️ Har bir kichik savol uchun alohida qatorda javob yozing:\n\n"
                f"Masalan:\n"
                f"javob1\n"
                f"javob2\n"
                f"javob3\n\n"
                f"⚠️ Barcha javoblarni majburiy ravishda kiriting!"
            )
            return

        except Exception as e:
            logger.error(f"42-savol javoblarini qayta ishlash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")

    elif step == 'problem_43_answers':
        # 43-savol (masalaviy) uchun javoblar kiritildi
        problem_43_input = update.message.text
        if problem_43_input.strip() == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        try:
            # Javoblarni qatorlarga ajratish
            lines = problem_43_input.split('\n')
            problem_43_answers = []

            for line in lines:
                line = line.strip()
                if line:  # Faqat bo'sh bo'lmagan qatorlarni olish
                    problem_43_answers.append(line)

            # Hech bo'lmaganda 1 ta javob bo'lishi kerak
            if len(problem_43_answers) == 0:
                await update.message.reply_text(
                    f"❌ 43-savol uchun hech bo'lmaganda 1 ta javob kiritish kerak!\n\n"
                    f"Qaytadan urinib ko'ring."
                )
                return

            # 43-savol javoblarini saqlash
            context.user_data['problem_43_answers'] = problem_43_answers

            # Javoblarni saqlash va testni yaratish
            # Savollarni yaratish
            questions = []

            # 1-35 savollar uchun ko'p tanlov savollarini yaratish
            mc_answers = context.user_data.get('mc_answers', [])
            for idx, answer in enumerate(mc_answers):
                if idx in [32, 33, 34]:  # 33, 34, 35-savollar
                    options = ['a', 'b', 'c', 'd', 'e', 'f']
                    questions.append({
                        'question': f"Savol {idx + 1}",
                        'options': options,
                        'correct': answer
                    })
                else:
                    options = ['a', 'b', 'c', 'd']
                    questions.append({
                        'question': f"Savol {idx + 1}",
                        'options': options,
                        'correct': answer
                    })

            # 36-40 savollar uchun yozma javob savollarini yaratish
            text_answers = context.user_data.get('text_answers', [])
            for idx, answer in enumerate(text_answers):
                question_idx = 35 + idx
                questions.append({
                    'question': f"Savol {question_idx + 1}",
                    'type': 'text_answer',
                    'options': [],
                    'correct': answer
                })

            # 41-savol (masalaviy) uchun javoblarni qo'shish
            problem_41_answers = context.user_data.get('problem_41_answers', [])
            questions.append({
                'question': f"Savol 41 (masalaviy)",
                'type': 'problem',
                'options': [],
                'correct': problem_41_answers,  # List of answers
                'sub_question_count': len(problem_41_answers)
            })

            # 42-savol (masalaviy) uchun javoblarni qo'shish
            problem_42_answers = context.user_data.get('problem_42_answers', [])
            questions.append({
                'question': f"Savol 42 (masalaviy)",
                'type': 'problem',
                'options': [],
                'correct': problem_42_answers,
                'sub_question_count': len(problem_42_answers)
            })

            # 43-savol (masalaviy) uchun javoblarni qo'shish
            problem_43_answers = context.user_data.get('problem_43_answers', [])
            questions.append({
                'question': f"Savol 43 (masalaviy)",
                'type': 'problem',
                'options': [],
                'correct': problem_43_answers,
                'sub_question_count': len(problem_43_answers)
            })

            # Barcha javoblar to'plandi - testni yaratish
            context.user_data['test_questions'] = questions

            # Test yaratish
            await save_test_immediately(update, context)
            return

        except Exception as e:
            logger.error(f"43-savol javoblarini qayta ishlash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")

    elif step == 'answers':
        # 1-35 savollar uchun javoblar kiritildi
        answers_text = update.message.text.strip()
        if answers_text.lower() == '/cancel':
            context.user_data.clear()
            await update.message.reply_text("❌ Test yaratish bekor qilindi.")
            return

        try:
            # Javoblarni qayta ishlash
            mc_answers = {}  # {question_num: answer}

            # Raqam + harf kombinatsiyalarini topish (1a, 2b, 10c, ...)
            pattern = r'(\d+)([a-fA-F])'
            matches = re.findall(pattern, answers_text)
            for num_str, letter in matches:
                q_num = int(num_str)
                if q_num <= 35:
                    mc_answers[q_num] = letter.lower()

            # Agar raqam+harf formatida yetarli javob bo'lmasa, eski formatni sinab ko'rish
            if len(mc_answers) < 35:
                # Eski format: faqat harflar (abc...)
                answers = []
                for char in answers_text.lower():
                    question_idx = len(answers)
                    if question_idx >= 35:
                        break
                    if question_idx in [32, 33, 34]:  # 33, 34, 35-savollar
                        if char in 'abcdef':
                            answers.append(char)
                    else:
                        if char in 'abcd':
                            answers.append(char)

                if len(answers) == 35:
                    # Eski format ishladi
                    mc_answers = {i+1: answers[i] for i in range(len(answers))}
                else:
                    await update.message.reply_text(
                        f"❌ Javoblar soni noto'g'ri!\n\n"
                        f"⚠️ 1-35 savollar uchun 35 ta javob kerak.\n"
                        f"🔢 Hozirgi son: {len(answers)} ta\n\n"
                        f"Format: 1a2b3c4d... yoki abc... (35 ta javob)\n"
                        f"⚠️ 33, 34, 35-savollar uchun e va f javoblar ham mumkin!"
                    )
                    return

            # 1-35 savollar uchun javoblarni listga o'tkazish
            mc_answers_list = []
            for q_num in range(1, 36):
                if q_num in mc_answers:
                    mc_answers_list.append(mc_answers[q_num])
                else:
                    await update.message.reply_text(
                        f"❌ {q_num}-savol uchun javob topilmadi!\n\n"
                        f"Barcha 1-35 savollar uchun javoblar kerak.\n"
                        f"Format: 1a2b3c4d... yoki abc... (35 ta javob)"
                    )
                    return

            # Javoblarni saqlash va 36-40 savollar uchun javoblarni so'rash
            context.user_data['mc_answers'] = mc_answers_list
            context.user_data['test_creation_step'] = 'text_answers'

            await update.message.reply_text(
                f"✅ 1-35 savollar uchun javoblar qabul qilindi!\n\n"
                f"📝 Endi 36-40 savollar uchun yozma javoblarni kiriting:\n\n"
                f"Har bir savol uchun alohida qatorda javob yozing.\n\n"
                f"Masalan:\n"
                f"Javob 36-savolga\n"
                f"Javob 37-savolga\n"
                f"Javob 38-savolga\n"
                f"Javob 39-savolga\n"
                f"Javob 40-savolga"
            )
            return

        except Exception as e:
            logger.error(f"1-35 javoblar qayta ishlash xatosi: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")



async def process_test_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test faylini qayta ishlash"""
    # Test yaratish yoki tahrirlash rejimida bo'lishi kerak
    is_creating = context.user_data.get('creating_test')
    is_editing = context.user_data.get('editing_test')

    if not is_creating and not is_editing:
        return

    step = context.user_data.get('test_creation_step') or context.user_data.get('test_editing_step')
    if step != 'file':
        return

    try:
        user_id = update.effective_user.id

        # Faylni yuklash
        document = update.message.document
        if not document:
            await update.message.reply_text("❌ Fayl topilmadi. Qayta yuboring.")
            return

        file = await context.bot.get_file(document.file_id)

        # Fayl mazmunini o'qish
        file_content = await file.download_as_bytearray()
        text_content = file_content.decode('utf-8', errors='ignore')

        # Test faylini saqlash (savollarni avtomatik ajratmaslik)
        # Foydalanuvchi o'zi test faylini yuboradi, savollar sonini avtomatik aniqlashga harakat qilmaymiz
        # Faylni saqlash
        test_files_dir = "test_files"
        os.makedirs(test_files_dir, exist_ok=True)

        # Vaqtinchalik fayl nomi
        temp_file_name = f"temp_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        file_ext = os.path.splitext(document.file_name or 'test.txt')[1] or '.txt'
        temp_file_path = os.path.join(test_files_dir, f"{temp_file_name}{file_ext}")

        # Faylni diskga saqlash
        with open(temp_file_path, 'wb') as f:
            f.write(file_content)

        # Fayl ma'lumotlarini saqlash
        context.user_data['test_file_path'] = temp_file_path
        context.user_data['test_file_name'] = document.file_name or 'test.txt'

        # Savollarni avtomatik ajratmaslik - foydalanuvchi javoblarni kiritadi
        # Savollar sonini avtomatik aniqlashga harakat qilmaymiz

        # Agar tahrirlash rejimida bo'lsa
        if is_editing:
            test_id = context.user_data.get('editing_test_id')
            data = load_data()
            if test_id and test_id in data['tests']:
                # Eski faylni o'chirish (agar mavjud bo'lsa)
                old_file_path = data['tests'][test_id].get('file_path')
                if old_file_path and os.path.exists(old_file_path):
                    try:
                        os.remove(old_file_path)
                    except:
                        pass

                # Yangi faylni saqlash
                test_files_dir = "test_files"
                os.makedirs(test_files_dir, exist_ok=True)
                file_name = document.file_name or "test.txt"
                file_ext = os.path.splitext(file_name)[1] or '.txt'
                new_file_path = os.path.join(test_files_dir, f"{test_id}{file_ext}")

                # Faylni diskga saqlash
                with open(new_file_path, 'wb') as f:
                    f.write(file_content)

                # Testni yangilash (savollarni o'zgartirmaslik, faqat faylni yangilash)
                data['tests'][test_id]['file_path'] = new_file_path
                data['tests'][test_id]['file_name'] = file_name
                save_data(data)

                context.user_data.pop('editing_test', None)
                context.user_data.pop('editing_test_id', None)
                context.user_data.pop('test_editing_step', None)
                await update.message.reply_text("✅ Test fayli yangilandi!\n\nJavoblarni yangilash uchun testni qayta tahrirlang.")
                return

        # Test yaratish rejimi
        # Faylni doimiy saqlash
        test_files_dir = "test_files"
        os.makedirs(test_files_dir, exist_ok=True)

        file_name = document.file_name or "test.txt"
        # Fayl nomini test ID bilan biriktirish uchun vaqtinchalik saqlash
        temp_file_path = os.path.join(test_files_dir, f"temp_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file_name}")

        # Faylni diskka yuklash
        await file.download_to_drive(temp_file_path)

        # Fayl yo'lini saqlash
        context.user_data['test_file_path'] = temp_file_path
        context.user_data['test_file_name'] = file_name
        context.user_data['test_creation_step'] = 'answers'

        # Javoblar kiritishni so'rash (1-35 savollar uchun ko'p tanlov, 36-40 uchun yozma javob)
        await update.message.reply_text(
            "✅ Fayl yuklandi!\n\n"
            "📝 Avval 1-35 savollar uchun javoblarni kiriting:\n\n"
            "Format: 1a2b3c4d... yoki abc... (35 ta javob)\n\n"
            "⚠️ Javoblar soni aniq 35 ta bo'lishi kerak!\n"
            "ℹ️ 33, 34, 35-savollar uchun e va f javoblar ham mumkin!\n\n"
            "📝 Keyin 36-40 savollar uchun yozma javoblar so'raladi."
        )

    except Exception as e:
        logger.error(f"Fayl qayta ishlash xatosi: {e}")
        await update.message.reply_text(f"❌ Xatolik: {str(e)}")


async def list_tests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Testlar ro'yxati"""
    if not await check_subscription(update, context):
        return

    # Ism va familya tekshiruvi
    if not await check_user_name(update, context):
        await update.message.reply_text(
            "❌ Botdan foydalanish uchun ism va familyangizni kiriting.\n\n"
            "Iltimos, /start ni bosing va ism va familyangizni kiriting."
        )
        return

    data = load_data()
    if not data['tests']:
        await update.message.reply_text("❌ Hozircha testlar mavjud emas.")
        return

    user_id = update.effective_user.id
    is_boss = user_id == BOSS_ID
    is_admin = user_id in data["admins"]

    keyboard = []
    for test_id, test_data in data['tests'].items():
        # Test yaratgan foydalanuvchi yoki admin/boss bo'lsa, tahrirlash tugmasi qo'shish
        can_edit = (is_boss or is_admin) and test_data.get('created_by') == user_id

        if can_edit:
            # Test nomi va tahrirlash tugmasi
            keyboard.append([
                InlineKeyboardButton(
                    test_data['name'],
                    callback_data=f"start_test_{test_id}"
                ),
                InlineKeyboardButton(
                    "✏️ Tahrirlash",
                    callback_data=f"edit_test_{test_id}"
                )
            ])
        else:
            # Oddiy foydalanuvchilar uchun faqat test nomi
            keyboard.append([InlineKeyboardButton(
                test_data['name'],
                callback_data=f"start_test_{test_id}"
            )])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("📋 Mavjud testlar:", reply_markup=reply_markup)


async def edit_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    """Testni tahrirlash"""
    user_id = update.effective_user.id
    data = load_data()

    if test_id not in data['tests']:
        if update.callback_query:
            await update.callback_query.answer("❌ Test topilmadi!")
        return

    test = data['tests'][test_id]

    # Faqat test yaratgan foydalanuvchi yoki boss tahrirlashi mumkin
    if user_id != BOSS_ID and test.get('created_by') != user_id:
        if update.callback_query:
            await update.callback_query.answer("❌ Bu testni tahrirlash huquqingiz yo'q!")
        return

    # Test tahrirlash rejimini boshlash
    context.user_data['editing_test'] = True
    context.user_data['editing_test_id'] = test_id
    context.user_data['test_editing_step'] = 'name'

    keyboard = [
        [InlineKeyboardButton("📝 Test nomini o'zgartirish", callback_data=f"edit_name_{test_id}")],
        [InlineKeyboardButton("📄 Test faylini qayta yuklash", callback_data=f"edit_file_{test_id}")],
        [InlineKeyboardButton("✅ Javoblarni o'zgartirish", callback_data=f"edit_answers_{test_id}")],
        [InlineKeyboardButton("📊 Testni natijalash", callback_data=f"finalize_test_{test_id}")],
        [InlineKeyboardButton("📋 0-1 Matrix yuklab olish", callback_data=f"download_matrix_{test_id}")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_edit")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    text = f"✏️ Test tahrirlash: {test['name']}\n\n"
    text += f"Savollar soni: {len(test['questions'])}\n\n"
    text += "Nimani tahrirlamoqchisiz?"

    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)


async def start_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    """Testni boshlash"""
    if not await check_subscription(update, context):
        return

    data = load_data()
    if test_id not in data['tests']:
        if update.callback_query:
            await update.callback_query.answer("❌ Test topilmadi!")
        else:
            await update.message.reply_text("❌ Test topilmadi!")
        return

    test = data['tests'][test_id]
    user_id = update.effective_user.id

    # Test to'xtatilganligini tekshirish
    if test.get('finalized', False):
        error_text = "❌ Bu test allaqachon natijalangan va to'xtatilgan!\n\nYangi testni ishlash mumkin emas."
        if update.callback_query:
            await update.callback_query.answer(error_text, show_alert=True)
        else:
            await update.message.reply_text(error_text)
        return

    # Foydalanuvchi bu testni allaqachon ishlaganligini tekshirish
    user_results = data.get('user_results', {})
    for r_id, result in user_results.items():
        if result.get('user_id') == user_id and result.get('test_id') == test_id:
            # Foydalanuvchi bu testni allaqachon ishlagan
            error_text = "❌ Siz bu testni allaqachon ishlagansiz!\n\nHar bir testni faqat bir marta ishlash mumkin."
            if update.callback_query:
                await update.callback_query.answer(error_text, show_alert=True)
            else:
                await update.message.reply_text(error_text)
            return

    # Testni boshlash
    context.user_data[f'test_{test_id}'] = {
        'test_id': test_id,
        'answers': {},
        'started_at': datetime.now().isoformat(),
        'waiting_answers': True
    }

    # Agar test fayli mavjud bo'lsa, uni yuborish
    if 'file_path' in test and os.path.exists(test['file_path']):
        try:
            # Faylni yuborish
            file_name = test.get('file_name', 'test.txt')
            if update.callback_query:
                await update.callback_query.edit_message_text(f"📝 {test['name']}\n\nTest fayli yuborilmoqda...")
                with open(test['file_path'], 'rb') as f:
                    await update.callback_query.message.reply_document(
                        document=f,
                        filename=file_name
                    )
                # 36-40 savollar mavjudligini tekshirish
                has_text_questions = any(
                    q.get('type') == 'text_answer'
                    for q in test['questions']
                )
                instruction_text = "✅ Test fayli yuborildi!\n\n"
                if has_text_questions:
                    instruction_text += "📝 1-35 savollar uchun javoblarni kiriting: 1a2b3c4d...\n"
                    instruction_text += "⚠️ 36-40 savollar uchun keyinroq yozma javoblar so'raladi."
                else:
                    instruction_text += "Javoblarni kiriting: 1a2b3c4d..."
                await update.callback_query.message.reply_text(instruction_text)
            else:
                await update.message.reply_text(f"📝 {test['name']}")
                with open(test['file_path'], 'rb') as f:
                    await update.message.reply_document(
                        document=f,
                        filename=file_name
                    )
                # 36-40 savollar mavjudligini tekshirish
                has_text_questions = any(
                    q.get('type') == 'text_answer'
                    for q in test['questions']
                )
                instruction_text = "✅ Test fayli yuborildi!\n\n"
                if has_text_questions:
                    instruction_text += "📝 1-35 savollar uchun javoblarni kiriting: 1a2b3c4d...\n"
                    instruction_text += "⚠️ 36-40 savollar uchun keyinroq yozma javoblar so'raladi."
                else:
                    instruction_text += "Javoblarni kiriting: 1a2b3c4d..."
                await update.message.reply_text(instruction_text)
        except Exception as e:
            logger.error(f"Fayl yuborish xatosi: {e}")
            # Agar fayl yuborib bo'lmasa, oddiy matn ko'rsatish
            text = f"📝 {test['name']}\n\n"
            has_text_questions = False
            for idx, question in enumerate(test['questions'], 1):
                text += f"{idx}. {question['question']}\n"
                if question.get('type') == 'text_answer':
                    text += "   (Yozma javob)\n"
                    has_text_questions = True
                else:
                    for opt_idx, option in enumerate(question['options']):
                        letter = chr(97 + opt_idx)  # a, b, c, d
                        text += f"   {letter}) {option}\n"
                text += "\n"
            if has_text_questions:
                text += "📝 1-35 savollar uchun javoblarni kiriting: 1a2b3c4d...\n"
                text += "⚠️ 36-40 savollar uchun keyinroq yozma javoblar so'raladi."
            else:
                text += "Javoblarni kiriting: 1a2b3c4d..."

            if update.callback_query:
                await update.callback_query.edit_message_text(text)
            else:
                await update.message.reply_text(text)
    else:
        # Agar fayl mavjud bo'lmasa, oddiy matn ko'rsatish
        text = f"📝 {test['name']}\n\n"
        has_text_questions = False
        for idx, question in enumerate(test['questions'], 1):
            text += f"{idx}. {question['question']}\n"
            if question.get('type') == 'text_answer':
                text += "   (Yozma javob)\n"
                has_text_questions = True
            else:
                for opt_idx, option in enumerate(question['options']):
                    letter = chr(97 + opt_idx)  # a, b, c, d
                    text += f"   {letter}) {option}\n"
            text += "\n"

        if has_text_questions:
            text += "📝 1-35 savollar uchun javoblarni kiriting: 1a2b3c4d...\n"
            text += "⚠️ 36-40 savollar uchun keyinroq yozma javoblar so'raladi."
        else:
            text += "Javoblarni kiriting: 1a2b3c4d..."

        if update.callback_query:
            await update.callback_query.edit_message_text(text)
        else:
            await update.message.reply_text(text)


async def process_test_answers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test javoblarini qayta ishlash (1a2b3c4d... formatida va yozma javoblar)"""
    if not update.message or not update.message.text:
        return False

    # Qaysi test uchun javob kiritilayotganini topish
    test_data_key = None
    test_id = None
    for key in context.user_data.keys():
        if key.startswith('test_'):
            test_data = context.user_data[key]
            # Faqat dictionary bo'lsa va 'waiting_answers' mavjud bo'lsa
            if isinstance(test_data, dict) and test_data.get('waiting_answers'):
                test_data_key = key
                test_id = test_data.get('test_id')
                break

    if not test_data_key or not test_id:
        return False  # Test ishlash rejimida emas

    user_id = update.effective_user.id
    text = update.message.text.strip()

    data = load_data()
    if test_id not in data['tests']:
        await update.message.reply_text("❌ Test topilmadi!")
        return False

    test = data['tests'][test_id]

    # Masalaviy javoblar kiritilayotganini tekshirish (41-43 savollar uchun)
    waiting_problem_answers = context.user_data[test_data_key].get('waiting_problem_answers', False)
    
    if waiting_problem_answers:
        # Masalaviy javoblar kiritilmoqda (41-43 savollar)
        lines = text.split('\n')
        problem_answers = []
        
        for line in lines:
            line = line.strip()
            if line:
                problem_answers.append(line)
        
        # 41-43 savollar uchun kerakli javoblar sonini tekshirish
        problem_questions = [q for q in test['questions'] if q.get('type') == 'problem']
        total_problem_answers_needed = sum(q.get('sub_question_count', 0) for q in problem_questions)
        
        if len(problem_answers) != total_problem_answers_needed:
            await update.message.reply_text(
                f"❌ Masalaviy javoblar soni noto'g'ri!\n\n"
                f"41-43 savollar uchun {total_problem_answers_needed} ta javob kerak.\n"
                f"Hozirgi son: {len(problem_answers)} ta\n\n"
                f"Har bir javobni alohida qatorda yozing."
            )
            return True
        
        # Masalaviy javoblarni saqlash
        # 41-43 savollar uchun (indices 40-42)
        answer_idx = 0
        for q_idx, question in enumerate(test['questions']):
            if question.get('type') == 'problem':
                sub_count = question.get('sub_question_count', 0)
                # Har bir masalaviy savol uchun bir nechta javob bo'lishi mumkin
                # Ularni vergul bilan birlashtiramiz
                sub_answers = problem_answers[answer_idx:answer_idx + sub_count]
                combined_answer = ','.join(sub_answers)
                context.user_data[test_data_key]['answers'][str(q_idx)] = combined_answer
                answer_idx += sub_count
        
        # Testni yakunlash
        context.user_data[test_data_key]['waiting_answers'] = False
        context.user_data[test_data_key]['waiting_problem_answers'] = False
        await finish_test(update, context, test_id)
        return True
    
    # Yozma javoblar kiritilayotganini tekshirish (36-40 savollar uchun)
    waiting_text_answers = context.user_data[test_data_key].get('waiting_text_answers', False)

    if waiting_text_answers:
        # Yozma javoblar kiritilmoqda (36-40 savollar)
        # Javoblarni qatorlarga ajratish (har bir qator bitta savol uchun)
        # Qavs ichidagi matn alohida javob deb hisoblanmaydi
        lines = text.split('\n')
        text_answers = []

        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Qavs ichidagi matn javobning bir qismi bo'lib qoladi, alohida javob emas
            text_answers.append(line)

        # 36-40 savollar uchun 5 ta javob kerak
        text_question_indices = [35, 36, 37, 38, 39]  # 0-based index

        if len(text_answers) != len(text_question_indices):
            await update.message.reply_text(
                f"❌ Yozma javoblar soni noto'g'ri!\n\n"
                f"36-40 savollar uchun {len(text_question_indices)} ta javob kerak.\n"
                f"Hozirgi son: {len(text_answers)} ta\n\n"
                f"Har bir savol uchun alohida qatorda javob yozing.\n"
                f"Qavs ichida tushuntirish yozishingiz mumkin (qavs ichidagi matn alohida javob deb hisoblanmaydi).\n\n"
                f"Masalan:\n"
                f"Javob 36-savolga (tushuntirish)\n"
                f"Javob 37-savolga\n"
                f"Javob 38-savolga (qavs ichidagi matn)\n"
                f"Javob 39-savolga\n"
                f"Javob 40-savolga"
            )
            return True

        # Yozma javoblarni saqlash
        for idx, answer in enumerate(text_answers):
            question_idx = text_question_indices[idx]
            context.user_data[test_data_key]['answers'][str(question_idx)] = answer

        # 41-43 savollar mavjudligini tekshirish
        has_problem_questions = any(
            q.get('type') == 'problem'
            for q in test['questions']
        )

        # Agar 41-43 savollar mavjud bo'lsa, ular uchun javoblarni so'rash
        if has_problem_questions:
            context.user_data[test_data_key]['waiting_text_answers'] = False
            context.user_data[test_data_key]['waiting_problem_answers'] = True
            
            # 41-43 savollar uchun kerakli javoblar sonini hisoblash
            problem_questions = [q for q in test['questions'] if q.get('type') == 'problem']
            total_problem_answers = sum(q.get('sub_question_count', 0) for q in problem_questions)
            
            await update.message.reply_text(
                f"✅ 36-40 savollar uchun javoblar qabul qilindi!\n\n"
                f"📝 Endi 41-43 masalaviy savollar uchun javoblarni kiriting:\n\n"
                f"Jami {total_problem_answers} ta javob kerak.\n"
                f"Har bir javobni alohida qatorda yozing.\n\n"
                f"Masalan:\n"
                f"javob1\n"
                f"javob2\n"
                f"javob3\n"
                f"..."
            )
            return True
        else:
            # 41-43 savollar yo'q bo'lsa, testni yakunlash
            context.user_data[test_data_key]['waiting_answers'] = False
            context.user_data[test_data_key]['waiting_text_answers'] = False
            await finish_test(update, context, test_id)
            return True

    else:
        # Ko'p tanlov javoblari kiritilmoqda (1-35 savollar)
        text_lower = text.lower()

        # 36-40 savollar mavjudligini tekshirish
        has_text_questions = any(
            q.get('type') == 'text_answer'
            for q in test['questions']
        )

        # Javoblarni ajratish (1a2b3c4d... yoki abc...)
        # Faqat 1-35 savollar uchun (0-based: 0-34)
        answers = []
        for char in text_lower:
            char_lower = char.lower()
            # Hozirgi savol indexini aniqlash (mavjud javoblar soni)
            question_idx = len(answers)

            # Faqat 1-35 savollar uchun javob qabul qilamiz (0-based: 0-34)
            if question_idx >= 35:
                break

            # 33, 34, 35-savollar (idx 32, 33, 34) uchun e va f ham qabul qilamiz
            if question_idx in [32, 33, 34]:  # 33, 34, 35-savollar
                if char_lower in 'abcdef':
                    answers.append(char_lower)
            else:
                # Boshqa savollar uchun faqat a-d qabul qilamiz
                if char_lower in 'abcd':
                    answers.append(char_lower)

        # 1-35 savollar uchun 35 ta javob kerak
        required_mc_answers = 35
        if len(answers) != required_mc_answers:
            await update.message.reply_text(
                f"❌ Javoblar soni noto'g'ri!\n"
                f"1-35 savollar uchun {required_mc_answers} ta javob kerak.\n"
                f"Hozirgi son: {len(answers)} ta\n\n"
                f"Qayta kiriting: 1a2b3c4d... formatida\n"
                f"⚠️ 33, 34, 35-savollar uchun e va f javoblar ham mumkin!"
            )
            return True

        # Ko'p tanlov javoblarini saqlash
        for idx, answer in enumerate(answers):
            context.user_data[test_data_key]['answers'][str(idx)] = answer

        # Agar 36-40 savollar mavjud bo'lsa, yozma javoblarni so'rash
        if has_text_questions:
            context.user_data[test_data_key]['waiting_text_answers'] = True
            await update.message.reply_text(
                f"✅ 1-35 savollar uchun javoblar qabul qilindi!\n\n"
                f"📝 Endi 36-40 savollar uchun yozma javoblarni kiriting:\n\n"
                f"Har bir savol uchun alohida qatorda javob yozing.\n\n"
                f"Masalan:\n"
                f"Javob 36-savolga\n"
                f"Javob 37-savolga\n"
                f"Javob 38-savolga\n"
                f"Javob 39-savolga\n"
                f"Javob 40-savolga"
            )
            return True
        else:
            # 36-40 savollar yo'q bo'lsa, testni yakunlash
            context.user_data[test_data_key]['waiting_answers'] = False
            await finish_test(update, context, test_id)
            return True


async def finish_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    """Testni yakunlash va natijalarni hisoblash"""
    test_data_key = f'test_{test_id}'
    if test_data_key not in context.user_data:
        return

    data = load_data()
    test = data['tests'][test_id]
    user_answers = context.user_data[test_data_key]['answers']
    user_id = update.effective_user.id

    # To'g'ri javoblar soni
    correct = 0
    total = len(test['questions'])
    results = []

    for idx, question in enumerate(test['questions']):
        user_answer = user_answers.get(str(idx), '')

        # Yozma javoblar uchun avtomatik tekshirish (36-40 savollar)
        if question.get('type') == 'text_answer':
            correct_answer = question.get('correct', '')
            # Case-insensitive tekshirish
            is_correct = user_answer.strip().lower() == correct_answer.strip().lower()
            if is_correct:
                correct += 1
            results.append({
                'question': question['question'],
                'user_answer': user_answer,
                'correct_answer': correct_answer,
                'is_correct': is_correct,
                'type': 'text_answer'
            })
        # Masalaviy javoblar uchun avtomatik tekshirish (41-43 savollar)
        elif question.get('type') == 'problem':
            correct_answers = question.get('correct', [])
            # User answer vergul bilan ajratilgan
            user_answers_list = [a.strip() for a in user_answer.split(',')]
            
            # Har bir javobni tekshirish va alohida saqlash
            sub_results = []
            correct_count = 0
            for i, (user_ans, correct_ans) in enumerate(zip(user_answers_list, correct_answers)):
                is_sub_correct = user_ans.strip().lower() == correct_ans.strip().lower()
                if is_sub_correct:
                    correct_count += 1
                sub_results.append({
                    'sub_index': i + 1,
                    'user_answer': user_ans.strip(),
                    'correct_answer': correct_ans.strip(),
                    'is_correct': is_sub_correct
                })
            
            # Agar barcha javoblar to'g'ri bo'lsa
            is_correct = (correct_count == len(correct_answers))
            if is_correct:
                correct += 1
            
            results.append({
                'question': question['question'],
                'user_answer': user_answer,
                'correct_answer': ','.join(correct_answers) if isinstance(correct_answers, list) else correct_answers,
                'is_correct': is_correct,
                'type': 'problem',
                'sub_results': sub_results,
                'sub_question_count': len(correct_answers)
            })
        else:
            # Ko'p tanlov javoblari uchun avtomatik tekshirish (1-35 savollar)
            is_correct = user_answer == question['correct']
            if is_correct:
                correct += 1
            results.append({
                'question': question['question'],
                'user_answer': user_answer,
                'correct_answer': question['correct'],
                'is_correct': is_correct
            })

    # Natijalarni saqlash (lekin hozir ko'rsatmaymiz)
    result_id = f"result_{user_id}_{test_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    if 'user_results' not in data:
        data['user_results'] = {}
    data['user_results'][result_id] = {
        'user_id': user_id,
        'test_id': test_id,
        'test_name': test['name'],
        'correct': correct,
        'total': total,
        'percentage': (correct / total * 100) if total > 0 else 0,
        'results': results,
        'completed_at': datetime.now().isoformat()
    }
    save_data(data)

    # 0-1 Matrix yaratish va yangilash
    from utils import generate_response_matrix
    matrix_file_path_1_40, matrix_file_path_41_43, _ = generate_response_matrix(test_id, data)
    if matrix_file_path_1_40:
        # Matrix faylini test ma'lumotlariga saqlash
        if 'matrix_file' not in test:
            test['matrix_file'] = matrix_file_path_1_40
        if 'matrix_file_41_43' not in test and matrix_file_path_41_43:
            test['matrix_file_41_43'] = matrix_file_path_41_43
        data['tests'][test_id] = test
        save_data(data)

    # Faqat "javobingiz qabul qilindi" deb yuborish
    # Natijalar testni natijalash tugmasi bosilguncha ko'rsatilmaydi
    text = "✅ Javobingiz qabul qilindi!\n\nTest natijalari o'qituvchi tomonidan e'lon qilingandan keyin ko'rsatiladi."

    if update.callback_query:
        await update.callback_query.edit_message_text(text)
    else:
        await update.message.reply_text(text)

    # User data tozalash
    del context.user_data[test_data_key]


async def finalize_test(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    """Testni natijalash - barcha natijalarni to'plash va o'qituvchiga yuborish"""
    user_id = update.effective_user.id
    data = load_data()

    if test_id not in data['tests']:
        if update.callback_query:
            await update.callback_query.answer("❌ Test topilmadi!")
        return

    test = data['tests'][test_id]

    # Faqat test yaratgan foydalanuvchi yoki boss natijalashi mumkin
    if user_id != BOSS_ID and test.get('created_by') != user_id:
        if update.callback_query:
            await update.callback_query.answer("❌ Bu testni natijalash huquqingiz yo'q!")
        return

    # Barcha foydalanuvchi natijalarini to'plash
    all_results = [
        r for r_id, r in data.get('user_results', {}).items()
        if r.get('test_id') == test_id
    ]

    if not all_results:
        if update.callback_query:
            await update.callback_query.answer("❌ Bu test uchun hali natijalar yo'q!")
        return

    # Natijalarni tayyorlash
    teacher_id = test.get('created_by')

    # Barcha natijalarni to'plash va tartiblash
    finalized_results = []
    for result in all_results:
        finalized_results.append({
            'user_id': result['user_id'],
            'correct': result['correct'],
            'total': result['total'],
            'percentage': result['percentage'],
            'completed_at': result['completed_at']
        })

    # Foiz bo'yicha tartiblash (yuqoridan pastga)
    finalized_results.sort(key=lambda x: x.get('percentage', 0), reverse=True)

    # Umumiy statistika
    total_students = len(finalized_results)
    avg_percentage = sum(r['percentage'] for r in finalized_results) / total_students if total_students > 0 else 0

    # O'qituvchiga natijalar xabari
    text = f"📊 Test natijalari: {test['name']}\n\n"
    text += f"📈 Umumiy statistika:\n"
    text += f"   Jami ishtirokchilar: {total_students}\n"
    text += f"   O'rtacha foiz: {avg_percentage:.1f}%\n\n"
    text += f"📋 Natijalar (foiz bo'yicha):\n\n"

    for idx, result in enumerate(finalized_results[:20], 1):  # Top 20
        text += f"{idx}. Talabgor: {result['user_id']}\n"
        text += f"   {result['correct']}/{result['total']} ({result['percentage']:.1f}%)\n\n"

    if total_students > 20:
        text += f"... va yana {total_students - 20} ta natija\n"

    # Excel fayl yaratish (barcha natijalar uchun)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Natijalari"
        
        # Header qator
        headers = ['#', 'Talabgor', 'To\'g\'ri javoblar', 'Jami savollar', 'Foiz (%)', 'Vaqt']
        ws.append(headers)
        
        # Header qatorini formatlash
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
        
        # Ma'lumotlar qatorlari
        for idx, result in enumerate(finalized_results, 1):
            completed_time = datetime.fromisoformat(result['completed_at']).strftime('%Y-%m-%d %H:%M')
            row = [
                idx,
                result['user_id'],
                result['correct'],
                result['total'],
                round(result['percentage'], 2),
                completed_time
            ]
            ws.append(row)
        
        # Ustunlarni kengaytirish
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20
        
        # Ma'lumotlar qatorlarini formatlash
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Excel faylni saqlash
        results_dir = "final_results"
        os.makedirs(results_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        excel_file_path = os.path.join(results_dir, f"test_results_{test_id}_{timestamp}.xlsx")
        wb.save(excel_file_path)
        
        # O'qituvchiga yuborish
        if update.callback_query:
            await update.callback_query.edit_message_text(text)
        else:
            await update.message.reply_text(text)

        # Excel faylni yuborish
        with open(excel_file_path, 'rb') as excel_file:
            if update.callback_query:
                await update.callback_query.message.reply_document(
                    document=excel_file,
                    filename=f"test_results_{test_id}.xlsx",
                    caption=f"📊 Test natijalari: {test['name']}\n\n"
                            f"📈 Jami ishtirokchilar: {total_students}\n"
                            f"📊 O'rtacha foiz: {avg_percentage:.1f}%"
                )
            else:
                await update.message.reply_document(
                    document=excel_file,
                    filename=f"test_results_{test_id}.xlsx",
                    caption=f"📊 Test natijalari: {test['name']}\n\n"
                            f"📈 Jami ishtirokchilar: {total_students}\n"
                            f"📊 O'rtacha foiz: {avg_percentage:.1f}%"
                )
        
    except Exception as e:
        logger.error(f"Excel fayl yaratish xatosi: {e}")
        error_text = f"❌ Excel fayl yaratishda xatolik: {str(e)}"
        if update.callback_query:
            await update.callback_query.message.reply_text(error_text)
        else:
            await update.message.reply_text(error_text)

    # 2ta matrix faylini yaratish va yuborish
        from utils import generate_response_matrix
        matrix_file_path_1_40, matrix_file_path_41_43, _ = generate_response_matrix(test_id, data)
        
        if matrix_file_path_1_40 and matrix_file_path_41_43:
            # 1. Questions 1-40 faylini yuborish
            try:
                with open(matrix_file_path_1_40, 'rb') as f1:
                    if update.callback_query:
                        await update.callback_query.message.reply_document(
                            document=f1,
                            filename=f"matrix_1-40_{test_id}.xlsx",
                            caption=f"📋 0-1 Matrix (1-dars): {test['name']}\n\n"
                                    f"📊 1-40 savollar uchun matrix\n"
                                    f"Format: user_id, Q1, Q2, ..., Q40\n"
                                    f"0 = xato javob, 1 = to'g'ri javob"
                        )
                    else:
                        await update.message.reply_document(
                            document=f1,
                            filename=f"matrix_1-40_{test_id}.xlsx",
                            caption=f"📋 0-1 Matrix (1-dars): {test['name']}\n\n"
                                    f"📊 1-40 savollar uchun matrix\n"
                                    f"Format: user_id, Q1, Q2, ..., Q40\n"
                                    f"0 = xato javob, 1 = to'g'ri javob"
                        )
            except Exception as e:
                logger.error(f"Matrix 1-40 yuborish xatosi: {e}")
            
            # 2. Questions 41-43 faylini yuborish
            try:
                with open(matrix_file_path_41_43, 'rb') as f2:
                    if update.callback_query:
                        await update.callback_query.message.reply_document(
                            document=f2,
                            filename=f"matrix_41-43_{test_id}.xlsx",
                            caption=f"📋 0-1 Matrix (2-dars): {test['name']}\n\n"
                                    f"📊 41-43 savollar uchun batafsil matrix\n"
                                    f"Format: Talabgor, 41.1, 41.2, ..., 43.n\n"
                                    f"Har bir kichik savol uchun alohida ustun\n"
                                    f"0 = xato javob, 1 = to'g'ri javob"
                        )
                    else:
                        await update.message.reply_document(
                            document=f2,
                            filename=f"matrix_41-43_{test_id}.xlsx",
                            caption=f"📋 0-1 Matrix (2-dars): {test['name']}\n\n"
                                    f"📊 41-43 savollar uchun batafsil matrix\n"
                                    f"Format: Talabgor, 41.1, 41.2, ..., 43.n\n"
                                    f"Har bir kichik savol uchun alohida ustun\n"
                                    f"0 = xato javob, 1 = to'g'ri javob"
                        )
            except Exception as e:
                logger.error(f"Matrix 41-43 yuborish xatosi: {e}")

        # Testni to'xtatish (o'chirmaslik, faqat to'xtatish)
        # Testni ishlashni to'xtatish uchun 'finalized' flag qo'shamiz
        test['finalized'] = True
        test['finalized_at'] = datetime.now(UZBEKISTAN_TZ).isoformat()
        data['tests'][test_id] = test
        save_data(data)

        success_text = f"✅ Test muvaffaqiyatli natijalandi va to'xtatildi!"
        if update.callback_query:
            await update.callback_query.message.reply_text(success_text)
        else:
            await update.message.reply_text(success_text)

    except Exception as e:
        logger.error(f"Test natijalash xatosi: {e}")
        error_text = f"❌ Xatolik: {str(e)}"
        if update.callback_query:
            await update.callback_query.answer(error_text, show_alert=True)
        else:
            await update.message.reply_text(error_text)


async def download_matrix(update: Update, context: ContextTypes.DEFAULT_TYPE, test_id: str):
    """0-1 Matrix yuklab olish"""
    user_id = update.effective_user.id
    data = load_data()

    if test_id not in data['tests']:
        if update.callback_query:
            await update.callback_query.answer("❌ Test topilmadi!")
        return

    test = data['tests'][test_id]

    # Faqat test yaratgan foydalanuvchi yoki boss yuklab olishi mumkin
    if user_id != BOSS_ID and test.get('created_by') != user_id:
        if update.callback_query:
            await update.callback_query.answer("❌ Bu testni matrixini yuklab olish huquqingiz yo'q!")
        return

    # Matrix yaratish/yangilash - ikkita alohida fayl
    from utils import generate_response_matrix
    file_path_1_40, file_path_41_43, matrix_text = generate_response_matrix(test_id, data)

    if not file_path_1_40 or not file_path_41_43:
        if update.callback_query:
            await update.callback_query.answer("❌ Matrix yaratib bo'lmadi yoki hali natijalar yo'q!", show_alert=True)
        else:
            await update.message.reply_text("❌ Matrix yaratib bo'lmadi yoki hali natijalar yo'q!")
        return

    # Ikkita matrix faylini yuborish
    try:
        # 1. Questions 1-40 faylini yuborish
        with open(file_path_1_40, 'rb') as f1:
            if update.callback_query:
                await update.callback_query.message.reply_document(
                    document=f1,
                    filename=f"matrix_1-40_{test_id}.xlsx",
                    caption=f"📋 0-1 Matrix (1-dars): {test['name']}\n\n"
                            f"📊 1-40 savollar uchun matrix\n"
                            f"Format: user_id, Q1, Q2, ..., Q40\n"
                            f"0 = xato javob, 1 = to'g'ri javob"
                )
            else:
                await update.message.reply_document(
                    document=f1,
                    filename=f"matrix_1-40_{test_id}.xlsx",
                    caption=f"📋 0-1 Matrix (1-dars): {test['name']}\n\n"
                            f"📊 1-40 savollar uchun matrix\n"
                            f"Format: user_id, Q1, Q2, ..., Q40\n"
                            f"0 = xato javob, 1 = to'g'ri javob"
                )
        
        # 2. Questions 41-43 faylini yuborish
        with open(file_path_41_43, 'rb') as f2:
            if update.callback_query:
                await update.callback_query.message.reply_document(
                    document=f2,
                    filename=f"matrix_41-43_{test_id}.xlsx",
                    caption=f"📋 0-1 Matrix (2-dars): {test['name']}\n\n"
                            f"📊 41-43 savollar uchun batafsil matrix\n"
                            f"Format: Talabgor, 41.1, 41.2, ..., 43.n\n"
                            f"Har bir kichik savol uchun alohida ustun\n"
                            f"0 = xato javob, 1 = to'g'ri javob"
                )
                await update.callback_query.answer("✅ Ikkala matrix ham yuborildi!")
            else:
                await update.message.reply_document(
                    document=f2,
                    filename=f"matrix_41-43_{test_id}.xlsx",
                    caption=f"📋 0-1 Matrix (2-dars): {test['name']}\n\n"
                            f"📊 41-43 savollar uchun batafsil matrix\n"
                            f"Format: Talabgor, 41.1, 41.2, ..., 43.n\n"
                            f"Har bir kichik savol uchun alohida ustun\n"
                            f"0 = xato javob, 1 = to'g'ri javob"
                )
    except Exception as e:
        logger.error(f"Matrix yuborish xatosi: {e}")
        # Agar fayl yuborib bo'lmasa, matn sifatida yuborish
        if update.callback_query:
            await update.callback_query.message.reply_text(
                f"📋 0-1 Matrix: {test['name']}\n\n"
                f"```\n{matrix_text}\n```",
                parse_mode='Markdown'
            )
        else:
            await update.message.reply_text(
                f"📋 0-1 Matrix: {test['name']}\n\n"
                f"```\n{matrix_text}\n```",
                parse_mode='Markdown'
            )


async def my_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi natijalari"""
    if not await check_subscription(update, context):
        return

    # Ism va familya tekshiruvi
    if not await check_user_name(update, context):
        await update.message.reply_text(
            "❌ Botdan foydalanish uchun ism va familyangizni kiriting.\n\n"
            "Iltimos, /start ni bosing va ism va familyangizni kiriting."
        )
        return

    user_id = update.effective_user.id
    data = load_data()

    # Faqat natijalangan testlar natijalarini ko'rsatish
    # Test natijalangan bo'lsa, u data['tests'] dan o'chirilgan bo'ladi
    user_results = []
    for r_id, r in data.get('user_results', {}).items():
        if r['user_id'] == user_id:
            test_id = r.get('test_id')
            # Agar test hali mavjud bo'lsa (natijalanmagan), natijalarni ko'rsatmaymiz
            if test_id not in data.get('tests', {}):
                user_results.append(r)

    if not user_results:
        await update.message.reply_text("❌ Sizda hali test natijalari yo'q yoki testlar hali natijalanmagan.")
        return

    text = "📊 Mening natijalarim:\n\n"
    for result in sorted(user_results, key=lambda x: x['completed_at'], reverse=True)[:10]:
        text += f"📝 {result['test_name']}\n"
        text += f"   {result['correct']}/{result['total']} ({result['percentage']:.1f}%)\n\n"

    await update.message.reply_text(text)



async def show_statistics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Statistika ko'rsatish - qisqacha muhim ma'lumotlar (faqat adminlar uchun)"""
    user_id = update.effective_user.id
    data = load_data()
    
    # Faqat adminlar va boss uchun
    if user_id != BOSS_ID and user_id not in data.get("admins", []):
        await update.message.reply_text("❌ Bu funksiya faqat adminlar uchun!")
        return
    
    # Foydalanuvchilar soni
    total_users = len(data.get('users', {}))
    
    # Testlar soni (faqat natijalanmagan testlar)
    active_tests = len([t for t_id, t in data.get('tests', {}).items() if not t.get('finalized', False)])
    finalized_tests = len([t for t_id, t in data.get('tests', {}).items() if t.get('finalized', False)])
    total_tests = active_tests + finalized_tests
    
    # Test topshirganlar soni (barcha natijalar)
    total_results = len(data.get('user_results', {}))
    
    # Bugungi test topshirganlar soni
    today = datetime.now(UZBEKISTAN_TZ).date()
    today_results = sum(1 for r in data.get('user_results', {}).values() 
                       if datetime.fromisoformat(r.get('completed_at', '2000-01-01')).date() == today)
    
    # Eng ko'p test topshirgan foydalanuvchi
    user_test_counts = {}
    for r in data.get('user_results', {}).values():
        uid = r.get('user_id')
        user_test_counts[uid] = user_test_counts.get(uid, 0) + 1
    
    top_user_id = max(user_test_counts.items(), key=lambda x: x[1])[0] if user_test_counts else None
    top_user_count = user_test_counts.get(top_user_id, 0) if top_user_id else 0
    
    # O'rtacha foiz (barcha natijalar uchun)
    all_percentages = [r.get('percentage', 0) for r in data.get('user_results', {}).values()]
    avg_percentage = sum(all_percentages) / len(all_percentages) if all_percentages else 0
    
    # Qisqacha statistika matni
    text = f"""📈 <b>Bot Statistika</b>

👥 <b>Jami foydalanuvchilar:</b> {total_users}
📝 <b>Faol testlar:</b> {active_tests}
✅ <b>Jami test topshirganlar:</b> {total_results}
📅 <b>Bugungi topshirganlar:</b> {today_results}
📊 <b>O'rtacha foiz:</b> {avg_percentage:.1f}%

<b>🏆 Eng faol foydalanuvchi:</b> {top_user_id} ({top_user_count} ta test)
"""
    
    # Reply keyboard yaratish (adminlar uchun to'liq keyboard)
    is_boss = user_id == BOSS_ID
    is_admin = user_id in data.get("admins", [])
    
    keyboard = [
        [KeyboardButton("📝 Test ishlash"), KeyboardButton("📊 Test natijalarim")]
    ]
    # Faqat adminlar va boss uchun qo'shimcha tugmalar
    if is_boss or is_admin:
        keyboard.append([KeyboardButton("➕ Test yaratish"), KeyboardButton("📈 Statistika")])
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    await update.message.reply_text(text, parse_mode='HTML', reply_markup=reply_markup)

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback query handler"""
    query = update.callback_query
    await query.answer()

    data = query.data

    # Admin boshqaruvi
    if data == "add_admin":
        await query.edit_message_text("Admin ID ni yuboring:")
        context.user_data['adding_admin'] = True
    elif data == "remove_admin":
        await query.edit_message_text("Olib tashlash kerak bo'lgan admin ID ni yuboring:")
        context.user_data['removing_admin'] = True
    elif data == "list_admins":
        data_file = load_data()
        admins = data_file['admins']
        if admins:
            text = "📋 Adminlar ro'yxati:\n\n" + "\n".join([f"• {admin_id}" for admin_id in admins])
        else:
            text = "❌ Adminlar mavjud emas."
        await query.edit_message_text(text)

    # Kanal boshqaruvi
    elif data == "add_channel":
        await query.edit_message_text("Kanal username yoki ID ni yuboring (masalan: @channel yoki -1001234567890):")
        context.user_data['adding_channel'] = True
    elif data == "remove_channel":
        await query.edit_message_text("Olib tashlash kerak bo'lgan kanal username yoki ID ni yuboring:")
        context.user_data['removing_channel'] = True
    elif data == "list_channels":
        data_file = load_data()
        channels = data_file['mandatory_channels']
        if channels:
            text = "📋 Majburiy kanallar:\n\n" + "\n".join([f"• {ch}" for ch in channels])
        else:
            text = "❌ Majburiy kanallar mavjud emas."
        await query.edit_message_text(text)

    # Test boshlash
    elif data.startswith("start_test_"):
        test_id = data.replace("start_test_", "")
        await start_test(update, context, test_id)

    # Postni ko'rish (ulashish)
    elif data.startswith("view_post_"):
        test_id = data.replace("view_post_", "")
        data_file = load_data()
        if test_id in data_file['tests']:
            test_data = data_file['tests'][test_id]
            try:
                bot_info = await context.bot.get_me()
                bot_username = f"@{bot_info.username}" if bot_info.username else "Test Bot"
            except:
                bot_username = "Test Bot"

            post_text, reply_markup = generate_test_post(test_data, test_id, bot_username)
            # Postni yangi xabar sifatida yuborish (nusxalab yuborish uchun)
            await query.answer("📢 Post yuborilmoqda...")
            await query.message.reply_text(
                post_text,
                parse_mode='HTML',
                reply_markup=reply_markup
            )
        else:
            await query.answer("❌ Test topilmadi!", show_alert=True)

    # Barcha testlar ro'yxati
    elif data == "list_all_tests":
        await query.answer("📋 Testlar ro'yxati yuborilmoqda...")
        # list_tests funksiyasini to'g'ridan-to'g'ri chaqirish
        # Lekin u message kerak, shuning uchun callback query message orqali yuboramiz
        data_file = load_data()
        if not data_file['tests']:
            await query.message.reply_text("❌ Hozircha testlar mavjud emas.")
            return

        user_id = query.from_user.id
        is_boss = user_id == BOSS_ID
        is_admin = user_id in data_file.get("admins", [])

        keyboard = []
        for test_id, test_data in data_file['tests'].items():
            can_edit = (is_boss or is_admin) and test_data.get('created_by') == user_id

            if can_edit:
                keyboard.append([
                    InlineKeyboardButton(
                        f"📝 {test_data['name']}",
                        callback_data=f"start_test_{test_id}"
                    ),
                    InlineKeyboardButton("✏️", callback_data=f"edit_test_{test_id}")
                ])
            else:
                keyboard.append([
                    InlineKeyboardButton(
                        f"📝 {test_data['name']}",
                        callback_data=f"start_test_{test_id}"
                    )
                ])

        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text(
            "📋 <b>Mavjud testlar:</b>\n\nTestni tanlang:",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )

    # Test tahrirlash
    elif data.startswith("edit_test_"):
        test_id = data.replace("edit_test_", "")
        await edit_test(update, context, test_id)

    # Test tahrirlash bosqichlari
    elif data.startswith("edit_name_"):
        test_id = data.replace("edit_name_", "")
        context.user_data['editing_test'] = True
        context.user_data['editing_test_id'] = test_id
        context.user_data['test_editing_step'] = 'name'
        await query.edit_message_text("Yangi test nomini kiriting:")

    elif data.startswith("edit_file_"):
        test_id = data.replace("edit_file_", "")
        context.user_data['editing_test'] = True
        context.user_data['editing_test_id'] = test_id
        context.user_data['test_editing_step'] = 'file'
        await query.edit_message_text("Yangi test faylini yuboring:")

    elif data.startswith("edit_answers_"):
        test_id = data.replace("edit_answers_", "")
        data_file = load_data()
        if test_id in data_file['tests']:
            test = data_file['tests'][test_id]
            context.user_data['editing_test'] = True
            context.user_data['editing_test_id'] = test_id
            context.user_data['test_editing_step'] = 'answers'
            context.user_data['test_questions'] = test['questions']
            await query.edit_message_text(f"Javoblarni kiriting: 1a2b3c4d...\n\nSavollar soni: {len(test['questions'])}")
        else:
            await query.answer("❌ Test topilmadi!")

    elif data.startswith("finalize_test_"):
        test_id = data.replace("finalize_test_", "")
        await finalize_test(update, context, test_id)

    elif data.startswith("download_matrix_"):
        test_id = data.replace("download_matrix_", "")
        await download_matrix(update, context, test_id)

    elif data == "cancel_edit":
        context.user_data.pop('editing_test', None)
        context.user_data.pop('editing_test_id', None)
        context.user_data.pop('test_editing_step', None)
        await query.edit_message_text("❌ Tahrirlash bekor qilindi.")


async def process_admin_channel_commands(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin va kanal qo'shish/olib tashlash"""
    if update.effective_user.id != BOSS_ID:
        return

    # Agar hech qanday operatsiya kutilayotgan bo'lmasa, hech narsa qilmaymiz
    if not any([
        context.user_data.get('adding_admin'),
        context.user_data.get('removing_admin'),
        context.user_data.get('adding_channel'),
        context.user_data.get('removing_channel')
    ]):
        return

    text = update.message.text.strip()
    data = load_data()

    # Admin qo'shish
    if context.user_data.get('adding_admin'):
        try:
            admin_id = int(text)
            if admin_id not in data['admins']:
                data['admins'].append(admin_id)
                save_data(data)
                await update.message.reply_text(f"✅ Admin {admin_id} qo'shildi!")
            else:
                await update.message.reply_text(f"⚠️ Bu admin allaqachon mavjud.")
            context.user_data['adding_admin'] = False
        except ValueError:
            await update.message.reply_text("❌ Noto'g'ri ID format. Faqat raqam kiriting.")

    # Admin olib tashlash
    elif context.user_data.get('removing_admin'):
        try:
            admin_id = int(text)
            if admin_id in data['admins']:
                data['admins'].remove(admin_id)
                save_data(data)
                await update.message.reply_text(f"✅ Admin {admin_id} olib tashlandi!")
            else:
                await update.message.reply_text(f"❌ Bu admin topilmadi.")
            context.user_data['removing_admin'] = False
        except ValueError:
            await update.message.reply_text("❌ Noto'g'ri ID format. Faqat raqam kiriting.")

    # Kanal qo'shish
    elif context.user_data.get('adding_channel'):
        channel = text.replace('@', '').strip()
        if channel not in data['mandatory_channels']:
            data['mandatory_channels'].append(channel)
            save_data(data)
            await update.message.reply_text(f"✅ Kanal {channel} qo'shildi!")
        else:
            await update.message.reply_text(f"⚠️ Bu kanal allaqachon mavjud.")
        context.user_data['adding_channel'] = False

    # Kanal olib tashlash
    elif context.user_data.get('removing_channel'):
        channel = text.replace('@', '').strip()
        if channel in data['mandatory_channels']:
            data['mandatory_channels'].remove(channel)
            save_data(data)
            await update.message.reply_text(f"✅ Kanal {channel} olib tashlandi!")
        else:
            await update.message.reply_text(f"❌ Bu kanal topilmadi.")
        context.user_data['removing_channel'] = False

